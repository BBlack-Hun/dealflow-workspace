"""투자사 딜공유 명단 가져오기 — **시트 두 장이 한 사람으로 합쳐지고, 두 번
넣어도 두 줄이 되지 않게.**

이 시트가 지금까지의 임포터로 안 읽히던 이유는 셋이다.

  1. **머리글이 두 줄**이다. 윗줄은 `8월`·`7월` 같은 달 묶음(세 칸 병합)이고
     아랫줄이 진짜 머리글이다. 윗줄을 안 읽으면 `딜소개` 라는 같은 이름의 칸이
     달 수만큼 생겨 **한 칸으로 뭉개진다.**
  2. 머리글이 **엑셀 수식**이다(`="딜소개 8/5 ("&COUNTIF(…)&")"`). 그대로 쓰면
     수식이 화면에 나오고, 세어 본 수는 열 때마다 달라져 같은 칸이 달마다
     **새 칸**으로 선다.
  3. `NO` 칸이 없다. 사람이 주인공이라 `이름` + `투자사명` 이 표를 연다.

그리고 이 명단만의 규칙이 둘 더 있다.

  · **번호가 없어도 넣는다.** 스타트업 임포터는 안 넣는데 여기는 넣는다(아래
    검사의 주석에 이유를 적었다).
  · **번호는 발송의 열쇠가 아니다.** 딜 소개는 이미 만들어진 카톡방으로 나가므로
    번호가 없어도 방 이름만 있으면 나간다.

이름·회사·번호는 전부 지어낸 값이다 — 저장소가 공개다.
"""
from __future__ import annotations

import csv
import re
from pathlib import Path

from datetime import date

import pytest

ROOT = Path(__file__).resolve().parent.parent

# 원본 시트가 그렇듯 이름에 괄호와 숫자가 붙는다. **코드가 이 이름을 알면 안 된다.**
LIST = "샘플 딜공유(9)"
OTHER = "샘플 다른명단(3)"

# ── 탭 ①: 명함. 머리글 한 줄, 번호가 여기에만 있다 ──────────────────────────
CARD_HEAD = ["이름", "관심도 (월말기준)", "카톡방 참여여부", "휴대폰", "회사",
             "전자 메일 주소", "부서", "직함"]

# ── 탭 ②: 딜공유. 머리글 **두 줄** + 수식 머리글 ────────────────────────────
#
# 윗줄은 세 칸을 병합해 단 달 묶음이라, 읽으면 맨 왼쪽 칸에만 값이 있고 나머지는
# 빈칸으로 온다. 원본이 그렇게 오므로 그대로 만든다.
DEAL_GROUP = ["", "", "", "", "8월", "", "", "7월", "", ""]
DEAL_HEAD = ["그룹/투자분야/라운드사이즈", "이름", "투자사명", "기타",
             # 사람이 쓴 말 + **세어 본 수**. 원본은 이 자리가 수식이고, 엑셀은
             # 계산해 둔 값을 함께 저장해서 앱은 이 모양으로 읽는다.
             "딜소개 8/5 (7) 8/12 (9)", "IR 요청 (투자사 2 / IR 3)",
             "미팅확정(1) 미팅완료(0)",
             "딜소개", "IR 요청 (투자사 0 / IR 0)", "미팅확정/미팅완료"]


def card_row(name, phone, firm, joined="●") -> list:
    return [name, "높음", joined, phone, firm, f"{name}@example.com", "투자본부", "이사"]


def deal_row(name, firm, aug="", memo="") -> list:
    return ["Pre IPO", name, firm, memo, aug, "", "", "", "", ""]


def book(tmp_path: Path, cards, deals, name="book.xlsx") -> str:
    """탭 두 장짜리 엑셀. 원본 워크북과 **같은 모양**으로 만든다."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "명함탭"
    ws.append(CARD_HEAD)
    for row in cards:
        ws.append(row)
    ws2 = wb.create_sheet("딜공유탭")
    ws2.append(DEAL_GROUP)
    ws2.append(DEAL_HEAD)
    for row in deals:
        ws2.append(row)
    path = tmp_path / name
    wb.save(path)
    wb.close()
    return str(path)


def run(monkeypatch, path, sheet, owner, *extra) -> int:
    """스크립트를 부르는 방식 그대로 부른다 — 인자까지가 이 도구의 규약이다.

    `--mode` 를 여기서 넣지 않는다. **부르는 사람이 고르는 값**이고(기본값을
    두면 위험한 쪽이 조용히 기본이 된다), 검사마다 어느 모드를 보는지가 곧
    그 검사의 내용이다.
    """
    import scripts.import_investor_list as tool

    argv = ["import_investor_list.py", path, "--sheet", sheet, "--owner", owner,
            *extra]
    monkeypatch.setattr("sys.argv", argv)
    return tool.main()


def run_book(monkeypatch, path, sheet, owner, *extra) -> int:
    """두 탭을 **한 명단으로** 합쳐 넣는다(만들기). 먼저 적은 탭이 이긴다."""
    return run(monkeypatch, path, sheet, owner, "--mode", "create",
               "--tab", "명함탭", "--tab", "딜공유탭", *extra)


def fill_book(monkeypatch, path, sheet, owner, *extra) -> int:
    """같은 두 탭을 **채우기**로 넣는다 — 줄도 칸도 만들지 않는다."""
    return run(monkeypatch, path, sheet, owner, "--mode", "fill",
               "--tab", "명함탭", "--tab", "딜공유탭", *extra)


def rows_in(db, label: str) -> list:
    from app.models import VcContact
    from app.services import sheet_owner

    db.expire_all()
    return [c for c in db.query(VcContact).all()
            if label in sheet_owner.labels_of(c.source_sheet)]


def month_values(db, contact, label: str) -> dict:
    """{칸 이름: 그 사람의 값}. 칸 id 를 키로 담긴 값을 이름으로 되돌린다."""
    from app.services import contact_columns as cc

    values = cc.load_notes(contact.notes)
    return {col.label: values.get(cc.note_key(col.id), "")
            for col in cc.month_columns(db, label, today=date(2026, 8, 15))}


@pytest.fixture()
def owners(db, users):
    return {"a": users["u1"].phone, "b": users["u2"].phone}


@pytest.fixture()
def sample(tmp_path):
    """한 사람이 두 탭에 나뉘어 있고, 각 탭에만 있는 사람이 하나씩."""
    return book(tmp_path,
                cards=[card_row("김샘플", "010-7000-0001", "샘플투자"),
                       card_row("이샘플", "010-7000-0002", "샘플파트너스")],
                deals=[deal_row("김샘플", "샘플투자", "8/5 샘플가/샘플나"),
                       deal_row("박샘플", "샘플캐피탈", "8/12 샘플다")])


# ── 1. 두 번 넣어도 두 배가 되지 않는다 ─────────────────────────────────────

def test_같은_파일을_두_번_넣어도_줄이_두_배가_되지_않는다(
        monkeypatch, db, owners, sample, capsys):
    """**이 파일에서 가장 중요한 검사다.**

    시트는 한 번에 다 맞는 일이 없다 — 다음 달 칸이 붙으면 또 올린다. 그때마다
    줄이 두 배가 되면 딜 소개가 두 번 나가고, 나간 뒤에는 되돌릴 수 없다.

    **번호가 없는 사람도 두 줄이 되면 안 된다.** 이쪽은 번호로 못 찾으므로
    이 명단 안에서 이름+투자사명으로 찾는다.
    """
    assert run_book(monkeypatch, sample, LIST, owners["a"], "--apply") == 0
    first = rows_in(db, LIST)
    assert len(first) == 3          # 김(두 탭) · 이(명함만) · 박(딜공유만)

    assert run_book(monkeypatch, sample, LIST, owners["a"], "--apply") == 0
    again = rows_in(db, LIST)
    assert len(again) == 3, (
        f"두 번 넣었더니 {len(first)}줄이 {len(again)}줄이 됐습니다 "
        "— 같은 사람에게 딜 소개가 두 번 나갑니다")
    assert {c.id for c in first} == {c.id for c in again}, "줄이 새로 만들어졌습니다"

    capsys.readouterr()
    run_book(monkeypatch, sample, LIST, owners["a"])
    assert re.search(r"새로 만들 줄\s+0\b", capsys.readouterr().out)


def test_두_번째_판은_앱에서_고친_값을_지우지_않는다(monkeypatch, db, owners, sample):
    """시트에서 비어 있는 칸이 앱의 값을 덮으면 임포트 한 번에 기록이 사라진다."""
    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    kept = [c for c in rows_in(db, LIST) if c.name == "박샘플"][0]
    kept.kakao_room_name = "박샘플 방"
    db.commit()

    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    db.expire_all()
    assert [c for c in rows_in(db, LIST) if c.name == "박샘플"][0] \
        .kakao_room_name == "박샘플 방"


def test_다시_넣어도_칸이_두_벌이_되지_않는다(monkeypatch, db, owners, sample):
    """칸이 두 벌이 되면 8월 기록이 두 칸에 갈려 어느 쪽이 최신인지 알 수 없다."""
    from app.services import contact_columns as cc

    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    db.expire_all()
    before = [c.label for c in cc.month_columns(db, LIST, today=date(2026, 8, 15))]
    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    db.expire_all()
    assert [c.label for c in cc.month_columns(db, LIST, today=date(2026, 8, 15))] == before


# ── 2. 번호가 없어도 들어간다 ───────────────────────────────────────────────

def test_번호가_없는_사람도_들어간다(monkeypatch, db, owners, sample, capsys):
    """**스타트업 임포터와 반대다.** 거기서는 번호가 없으면 건너뛴다.

    저쪽 사정: 앱에 이미 서른두 줄이 들어와 있어서, 번호가 없으면 그 줄과 같은
    사람인지 가릴 수가 없었다. 조용히 넣으면 그게 두 줄이 된다.

    이쪽 사정: **명단 자체가 새로 생긴다.** 이 명단 안에는 아직 겹칠 상대가
    없다. 그리고 이 시트에는 번호 칸이 아예 없어서(원본 워크북 전 탭을 뒤져도
    없다) 안 넣으면 그 사람들이 통째로 사라진다 — 명단도 이력도 화면에 안 뜬다.

    번호를 버려도 되는 이유가 하나 더 있다. **번호는 발송의 열쇠가 아니다.**
    딜 소개는 이미 만들어진 카톡방으로 나가고(`dashboard._room_state`),
    번호의 쓸모는 사람을 대조하는 것이다.
    """
    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")

    got = {c.name: c for c in rows_in(db, LIST)}
    assert "박샘플" in got, "번호 없는 사람이 빠졌습니다 — 명단에서 통째로 사라집니다"
    assert not (got["박샘플"].phone or "").strip()
    # 몇 명이 그런지 **적어서 알린다.** 나중에 번호를 채울 때 확인해야 할 목록이다.
    assert "번호 없이 들어가는 사람" in capsys.readouterr().out


# ── 3. 발송은 번호가 아니라 **카톡방**으로 나간다 ───────────────────────────

def test_번호가_없어도_카톡방_이름이_있으면_발송_대상이다(
        monkeypatch, db, owners, sample):
    """예전에 "번호가 없으면 발송이 안 된다" 고 적어 두었는데 **틀린 말이었다.**

    발송 준비 상태를 정하는 곳(`dashboard._room_state`)은 채널이 카톡인가 ·
    방 이름이 있는가 · 확인됐는가만 본다. **번호는 조건에 없다.**
    """
    from app.services.dashboard import _room_state

    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    who = [c for c in rows_in(db, LIST) if c.name == "박샘플"][0]

    assert _room_state(who) == "missing", "방이 없으니 `방 미등록` 이어야 합니다"
    who.kakao_room_name = "박샘플 방"
    db.commit()
    assert _room_state(who) in ("verified", "unverified"), (
        "번호가 없어도 방 이름이 있으면 보낼 수 있습니다")


def test_방도_번호도_없으면_방_미등록으로_보인다(monkeypatch, db, owners, sample):
    """**감추지 않는다.** 고쳐야 할 것이 그대로 보이는 것이 정확한 상태다.

    `채널 불가 투자사`(= 보낼 길이 없다)로 뜨면 안 된다 — 달마다의 딜공유
    기록이 있는 사람은 이미 카톡방으로 받고 있고, 우리가 방 제목을 모를 뿐이다.
    """
    from app.services.dashboard import ROOM_LABELS, _room_state

    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    states = {c.name: _room_state(c) for c in rows_in(db, LIST)}
    assert set(states.values()) == {"missing"}, states
    assert ROOM_LABELS["missing"][0] == "방 미등록"


def test_방_이름을_지어내지_않는다(monkeypatch, db, owners, sample):
    """지어 준 방 제목이 실제와 다르면 **발송이 통째로 skip 된다.**

    이 시트들의 실제 방 제목은 앱이 짓는 모양과 아예 다르다(대화 기록에 남은
    방 이름이 그렇게 말한다). 모르면 비워 두는 편이 낫다.
    """
    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    assert not any((c.kakao_room_name or "").strip() for c in rows_in(db, LIST))


# ── 4. 탭 두 장이 한 사람으로 합쳐진다 ──────────────────────────────────────

def test_연락처는_명함_탭에서_월별_이력은_딜공유_탭에서_온다(
        monkeypatch, db, owners, sample):
    """한 사람의 자료가 탭 두 장에 나뉘어 있다.

    따로 넣으면 같은 사람이 두 줄이 되고, 한쪽만 넣으면 연락처나 이력이 통째로
    빠진다. 합쳐야 한 사람이 온전해진다.
    """
    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    got = {c.name: c for c in rows_in(db, LIST)}

    assert len(got) == 3, "겹치는 사람이 두 줄이 됐습니다"
    both = got["김샘플"]
    assert both.phone == "010-7000-0001", "연락처가 명함 탭에서 안 왔습니다"
    assert both.email == "김샘플@example.com"
    assert month_values(db, both, LIST)["8월 딜소개 8/5 8/12"] \
        == "8/5 샘플가/샘플나", "월별 이력이 딜공유 탭에서 안 왔습니다"
    # 각 탭에만 있는 사람도 빠지지 않는다.
    assert (got["이샘플"].phone, got["박샘플"].phone) == ("010-7000-0002", None)


def test_먼저_적은_탭의_값이_이긴다(monkeypatch, db, owners, tmp_path):
    """탭 순서가 곧 우선순위다. 나중 탭이 앞 탭을 밀어내면 어느 탭을 먼저
    적었느냐에 따라 결과가 달라진다."""
    path = book(tmp_path,
                cards=[card_row("김샘플", "010-7000-0001", "샘플투자")],
                deals=[deal_row("김샘플", "샘플투자")])
    run_book(monkeypatch, path, LIST, owners["a"], "--apply")
    assert rows_in(db, LIST)[0].firm == "샘플투자"


# ── 5. 머리글 두 줄 ─────────────────────────────────────────────────────────

def test_머리글_윗줄의_달_묶음이_칸_이름에_들어간다(monkeypatch, db, owners, sample):
    """윗줄을 안 읽으면 `딜소개` 가 달 수만큼 생겨 **한 칸으로 뭉개진다.**

    붙이는 것은 없는 말을 지어내는 것이 아니라 **윗줄에 적힌 것을 제자리에
    돌려놓는 것**이다.
    """
    from app.services import contact_columns as cc

    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    db.expire_all()
    labels = [c.label for c in cc.month_columns(db, LIST, today=date(2026, 8, 15))]

    assert labels == ["8월 딜소개 8/5 8/12", "8월 IR 요청", "8월 미팅확정 미팅완료",
                      "7월 딜소개", "7월 IR 요청", "7월 미팅확정/미팅완료"], labels
    assert len(set(labels)) == len(labels), "같은 이름의 칸이 뭉개졌습니다"


def test_달_묶음은_고정_칸에서_끊긴다():
    """묶음은 세 칸 병합이라 오른쪽으로 이어받는데, **고정 칸을 만나면 끊는다.**

    안 끊으면 표 맨 끝의 `대화내역 메모` 가 마지막 달 것으로 읽히고, 윗줄 첫
    칸에 적어 둔 메모가 뒤따르는 칸 이름에 붙는다.
    """
    from scripts.import_investor_list import parse_tab

    rows = [["운영 메모", "", "", "8월", "", "", ""],
            ["이름", "투자사명", "기타", "딜소개", "IR 요청", "미팅", "대화내역 메모"],
            ["김샘플", "샘플투자", "", "8/5 샘플가", "", "", "통화함"]]
    got = parse_tab(rows)

    assert got["columns"] == ["8월 딜소개", "8월 IR 요청", "8월 미팅"], got["columns"]
    assert "대화내역 메모" not in " ".join(got["columns"]), (
        "표 끝의 칸이 마지막 달 것으로 읽혔습니다")
    assert got["items"][0]["fields"]["memo"] == "통화함"


# ── 6. 수식 문자열이 칸 이름으로 새어 나오지 않는다 ─────────────────────────

def test_수식_문자열이_칸_이름으로_새어_나오지_않는다(monkeypatch, db, owners, tmp_path):
    """계산값이 저장돼 있지 않은 파일·CSV 로 내보낸 파일에서는 **수식이 그대로**
    올라온다. 그대로 두면 `="딜소개 8/5 ("&COUNTIF(…` 가 화면의 칸 이름이 된다."""
    from app.services import contact_columns as cc

    path = tmp_path / "formula.csv"
    with path.open("w", encoding="utf-8", newline="") as fp:
        writer = csv.writer(fp)
        writer.writerow(["", "", "", "8월", "", ""])
        writer.writerow([
            "이름", "투자사명", "기타",
            '="딜소개 8/5 ("&COUNTIF(D3:D99,"*8/5*")&") 8/12 ("'
            '&COUNTIF(D3:D99,"*8/12*")&")"',
            '="IR 요청 (투자사 "&COUNTIF(E3:E99,"*전달*")&" / IR "'
            '&SUMPRODUCT((LEN(E3:E99)))&")"',
            '="미팅확정("&SUMPRODUCT((LEN(F3:F99)))&")"',
        ])
        writer.writerow(["김샘플", "샘플투자", "", "8/5 샘플가", "", ""])
    run(monkeypatch, str(path), LIST, owners["a"], "--mode", "create", "--apply")
    db.expire_all()

    labels = [c.label for c in cc.month_columns(db, LIST, today=date(2026, 8, 15))]
    assert labels == ["8월 딜소개 8/5 8/12", "8월 IR 요청", "8월 미팅확정"], labels
    for label in labels:
        for mark in ("=", "COUNTIF", "SUMPRODUCT", '"', "&"):
            assert mark not in label, f"수식이 칸 이름에 남았습니다: {label!r}"


def test_수식이든_계산된_값이든_칸_이름이_같아야_한다():
    """같은 칸을 두 가지 모양으로 읽는다 — 수식 그대로, 또는 엑셀이 계산해 둔 값.

    두 모양이 다른 이름이 되면 **한 파일은 새 칸을, 다른 파일은 옛 칸을** 채워
    같은 달의 기록이 두 칸에 갈린다.
    """
    from scripts.import_investor_list import clean_label

    pairs = [
        ('="딜소개 8/5 ("&COUNTIF(E3:E133,"*8/5*")&") 8/12 ("'
         '&COUNTIF(E3:E133,"*8/12*")&")"',
         "딜소개 8/5 (101) 8/12 (104)"),
        ('="IR 요청 (투자사 "&COUNTIF(F3:F133,"*전달*")&" / IR "'
         '&SUMPRODUCT(LEN(F3:F133))&")"',
         "IR 요청 (투자사 4 / IR 14)"),
        ('="미팅확정("&SUMPRODUCT(LEN(G3:G133))&") 미팅완료("'
         '&SUMPRODUCT(LEN(G3:G133))&")"',
         "미팅확정(8) 미팅완료(4)"),
    ]
    for formula, computed in pairs:
        assert clean_label(formula) == clean_label(computed), formula
    assert clean_label(pairs[1][0]) == "IR 요청"


def test_사람이_손으로_쓴_괄호는_남긴다():
    """세어 본 수만 뗀다. 사람이 쓴 말을 떼면 **뜻이 바뀐다.**"""
    from scripts.import_investor_list import clean_label

    assert clean_label("관심도 (월말기준)") == "관심도 (월말기준)"
    assert clean_label("딜소개\n(8/26 딜소개 없음)") == "딜소개 (8/26 딜소개 없음)"
    assert clean_label("공통 (9월~)") == "공통 (9월~)"
    # 줄바꿈은 공백으로 편다 — 머리글 한 칸에 두 줄로 들어 있다.
    assert clean_label("IR 요청 (투자사 29 / IR 66)\n/ 미팅 안내") == "IR 요청 / 미팅 안내"


# ── 앱 전체와는 **번호로만** 잇는다 ─────────────────────────────────────────

def test_앱에_이미_있는_번호는_새로_만들지_않고_주인만_바뀐다(
        monkeypatch, db, users, owners, tmp_path):
    """그 줄에 **카톡방과 발송 이력**이 붙어 있다. 옮기는 것이지 베끼는 것이 아니다."""
    from app.models import SheetOwner, VcContact

    db.add(SheetOwner(label=OTHER, user_id=users["u1"].id))
    db.add(VcContact(user_id=users["u1"].id, source_sheet=OTHER, name="최샘플",
                     firm="샘플에쿼티", phone="010-7000-0009", channel_kakao=1,
                     kakao_room_name="최샘플 방", room_verified="verified"))
    db.commit()
    before = db.query(VcContact).count()

    path = book(tmp_path,
                cards=[card_row("최샘플", "010-7000-0009", "샘플에쿼티(사명 변경)")],
                deals=[])
    run_book(monkeypatch, path, LIST, owners["b"], "--apply")
    db.expire_all()

    assert db.query(VcContact).count() == before, "줄이 새로 만들어졌습니다"
    moved = rows_in(db, LIST)[0]
    assert moved.user_id == users["u2"].id, "주인이 안 바뀌었습니다"
    assert moved.kakao_room_name == "최샘플 방", "카톡방 이력이 끊겼습니다"
    assert OTHER not in (moved.source_sheet or ""), (
        "옛 담당자의 명단에 그대로 남아 있습니다 — 딜 소개가 두 번 나갑니다")


def test_이름이_같아도_남의_명단_줄에_붙지_않는다(
        monkeypatch, db, users, owners, tmp_path):
    """**동명이인이 있다.** 한 이름이 셋인 적도 있었다.

    이름으로 앱 전체를 뒤지면 남의 카톡방으로 딜 소개가 나간다. 이름+투자사명은
    **이 명단 안에서만** 쓴다.
    """
    from app.models import SheetOwner, VcContact

    db.add(SheetOwner(label=OTHER, user_id=users["u1"].id))
    db.add(VcContact(user_id=users["u1"].id, source_sheet=OTHER, name="김샘플",
                     firm="샘플투자", phone="010-7000-0077",
                     kakao_room_name="남의 방"))
    db.commit()

    path = book(tmp_path, cards=[], deals=[deal_row("김샘플", "샘플투자")])
    run(monkeypatch, path, LIST, owners["b"], "--mode", "create",
        "--tab", "딜공유탭", "--apply")
    db.expire_all()

    made = rows_in(db, LIST)
    assert len(made) == 1
    assert made[0].kakao_room_name is None, (
        "이름이 같다는 이유로 남의 줄을 가져왔습니다 — 남의 방으로 딜 소개가 나갑니다")
    assert OTHER in (db.query(VcContact)
                     .filter(VcContact.phone == "010-7000-0077").one().source_sheet)


def test_배정표가_다른_명단으로_정한_사람은_안_들어간다(
        monkeypatch, db, owners, tmp_path, capsys):
    """겹치는 사람을 양쪽에 넣으면 두 계정에 들어가 딜 소개가 두 번 나간다."""
    path = book(tmp_path,
                cards=[card_row("김샘플", "010-7000-0001", "샘플투자"),
                       card_row("이샘플", "010-7000-0002", "샘플파트너스")],
                deals=[])
    rules = tmp_path / "rulings.csv"
    with rules.open("w", encoding="utf-8", newline="") as fp:
        csv.writer(fp).writerows([["# 번호", "넣을 명단"],
                                  ["010-7000-0002", OTHER]])
    run_book(monkeypatch, path, LIST, owners["a"], "--apply",
             "--rulings", str(rules))

    assert {c.name for c in rows_in(db, LIST)} == {"김샘플"}
    assert "다른 명단으로 배정됨" in capsys.readouterr().out


# ── 명단 설정 ───────────────────────────────────────────────────────────────

def test_새_명단은_투자사로_세고_딜공유_배치로_선다(
        monkeypatch, db, owners, sample):
    """**스타트업 명단과 반대다.** 이들은 진짜 투자사라 딜 소개를 받을 사람들이고,
    투자사 수와 발송 대상에 들어가야 맞다."""
    from app.models import SheetOwner
    from app.services import contact_columns as cc
    from app.services import sheet_owner

    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    db.expire_all()

    settings = db.query(SheetOwner).filter(SheetOwner.label == LIST).one()
    assert settings.layout == cc.INVESTOR_MONTHLY
    assert settings.is_hidden == 0
    assert len(sheet_owner.investors(db, rows_in(db, LIST))) == 3, (
        "투자사로 세지 않으면 딜소개 발송 대상에서 빠집니다")


def test_투자사로_세지_않기로_해_둔_것을_다시_올려도_되돌아가지_않는다(
        monkeypatch, db, owners, sample):
    """화면에서 사람이 정한 값을 임포트 한 번이 되돌리면 안 된다."""
    from app.models import SheetOwner

    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    db.expire_all()
    db.query(SheetOwner).filter(SheetOwner.label == LIST).one().is_hidden = 1
    db.commit()

    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    db.expire_all()
    assert db.query(SheetOwner).filter(SheetOwner.label == LIST).one().is_hidden == 1


def test_미리보기가_먼저다(monkeypatch, db, owners, sample, capsys):
    """`--apply` 없이는 한 줄도 안 들어간다."""
    run_book(monkeypatch, sample, LIST, owners["a"])

    assert rows_in(db, LIST) == []
    out = capsys.readouterr().out
    assert "새로 만들 줄" in out and "--apply" in out


def test_담당자_계정이_없으면_만들지_않고_멈춘다(
        monkeypatch, db, users, sample, capsys):
    """번호를 한 자 잘못 적었을 때 **유령 계정**이 생기고 거기에 사람들이 붙는다."""
    from app.models import User

    assert run_book(monkeypatch, sample, LIST, "010-9999-0000", "--apply") == 1
    assert db.query(User).count() == len(users)
    assert rows_in(db, LIST) == []
    assert "add_user.py" in capsys.readouterr().out


def test_투자사명이_비어도_사람을_잃지_않는다(monkeypatch, db, owners, tmp_path, capsys):
    """**임포트에서 사람을 잃는 것이 가장 나쁘다.**

    투자사명 칸이 비었을 뿐 그 줄에는 달마다의 기록도 메모도 들어 있다.
    다만 이름만으로 찾게 되니 몇 명이 그런지 적어서 알린다.
    """
    path = book(tmp_path, cards=[],
                deals=[deal_row("한샘플", "", "8/5 샘플가"),
                       deal_row("정샘플", "샘플벤처스")])
    run(monkeypatch, path, LIST, owners["a"], "--mode", "create",
        "--tab", "딜공유탭", "--apply")

    assert {c.name for c in rows_in(db, LIST)} == {"한샘플", "정샘플"}
    assert "투자사명이 빈 줄" in capsys.readouterr().out


# ── 공개 저장소 ─────────────────────────────────────────────────────────────

def test_명단_이름과_연락처가_스크립트에_박혀_있지_않다():
    """저장소가 공개다. 그리고 이름이 박혀 있으면 다음 명단에서 또 박아야 한다."""
    src = (ROOT / "scripts" / "import_investor_list.py").read_text(encoding="utf-8")
    src = re.sub(r'"""(?:.|\n)*?"""', "", src)
    src = re.sub(r"^\s*#.*$", "", src, flags=re.M)

    phones = re.findall(r"\b01\d[-\s]?\d{3,4}[-\s]?\d{4}\b", src)
    assert not phones, f"연락처가 코드에 박혀 있습니다: {phones}"
    for banned in (LIST, OTHER, "딜공유현황", "심사역 리스트"):
        assert banned not in src, f"명단 이름 `{banned}` 이 코드에 박혀 있습니다"


def test_칸이_한_칸씩_밀린_줄을_넣기_전에_알린다(monkeypatch, db, owners, tmp_path, capsys):
    """실제로 휴대폰 칸이 비고 회사·부서·직함이 통째로 오른쪽으로 밀린 줄이 있었다.

    **버리지 않는다** — 버리면 그 사람이 사라진다. 그런데 조용히 넣으면 더 나쁘다:
    탭을 합치는 열쇠가 (이름, 투자사명)이라, 한 탭에서는 번호가 다른 탭에서는
    진짜 투자사명이 열쇠가 되어 **같은 사람이 두 줄로 갈린다.**

    그래서 미리보기에서 알린다. 시트를 고친 뒤 넣는 것이 맞는 순서다.
    """
    path = book(tmp_path,
                # 회사 칸에 번호가 들어온 줄(한 칸씩 밀렸다)
                cards=[["이샘플", "낮음", "○", "", "010-7000-0055",
                        "샘플증권", "IB부문", "차장"]],
                deals=[deal_row("이샘플", "샘플증권", "8/5 샘플가")])
    run_book(monkeypatch, path, LIST, owners["a"])
    out = capsys.readouterr().out

    assert "밀린 것으로 보이는 줄" in out, "조용히 두 줄이 됩니다"
    assert "이샘플" in out
    assert rows_in(db, LIST) == [], "미리보기인데 들어갔습니다"


# ── 채우기 (`--mode fill`) ──────────────────────────────────────────────────
#
# 명단은 이미 서 있고, 나중에 다른 탭에서 **명함을 찾아 얹는** 일이다. 만들기와
# 섞으면 두 가지가 조용히 어긋난다 — 못 찾은 사람이 새 줄로 서서 같은 사람이 두
# 줄이 되고, 앱에서 고쳐 둔 값이 시트 한 장에 덮인다.

@pytest.fixture()
def standing(monkeypatch, db, owners, tmp_path):
    """딜공유 탭만으로 **먼저 세워 둔** 명단. 명함이 통째로 비어 있다.

    원본이 그랬다 — 딜공유 탭에는 번호 칸이 아예 없어서, 그 탭만 넣으면 이름·
    투자사명·달마다의 기록만 있는 줄이 선다.
    """
    path = book(tmp_path,
                cards=[card_row("김샘플", "010-7000-0001", "샘플투자"),
                       card_row("이샘플", "010-7000-0002", "샘플파트너스")],
                deals=[deal_row("김샘플", "샘플투자", "8/5 샘플가/샘플나"),
                       deal_row("이샘플", "샘플파트너스", "8/12 샘플다")],
                name="standing.xlsx")
    run(monkeypatch, path, LIST, owners["a"], "--mode", "create",
        "--tab", "딜공유탭", "--apply")
    db.expire_all()
    return path


def test_채우기는_줄을_만들지_않는다(monkeypatch, db, owners, standing, tmp_path, capsys):
    """**만들면 같은 사람이 두 줄이 되고, 그 순간 딜 소개가 두 번 나간다.**

    시트에 이 명단에 없는 사람이 섞여 있는 것은 흔한 일이다(명함 탭이 더 넓다).
    """
    wider = book(tmp_path,
                 cards=[card_row("김샘플", "010-7000-0001", "샘플투자"),
                        card_row("최샘플", "010-7000-0009", "샘플캐피탈")],
                 deals=[], name="wider.xlsx")
    before = {c.id for c in rows_in(db, LIST)}

    assert run(monkeypatch, wider, LIST, owners["a"], "--mode", "fill",
               "--tab", "명함탭", "--apply") == 0
    assert {c.id for c in rows_in(db, LIST)} == before, "채우기가 줄을 만들었습니다"
    assert "채우기 모드에서는 만들지 않는다" in capsys.readouterr().out


def test_채우기는_빈_칸에만_얹고_두_번_돌리면_0칸이다(
        monkeypatch, db, owners, standing, capsys):
    """앱에서 고쳐 둔 값을 시트 한 장이 덮으면 안 된다. 그리고 빈 칸만 채우면
    **두 번 돌았는지가 결과로 보인다.**"""
    kept = [c for c in rows_in(db, LIST) if c.name == "김샘플"][0]
    kept.title = "손으로 고친 직함"
    db.commit()

    fill_book(monkeypatch, standing, LIST, owners["a"], "--apply")
    first = capsys.readouterr().out
    db.expire_all()
    filled = [c for c in rows_in(db, LIST) if c.name == "김샘플"][0]
    assert filled.phone == "010-7000-0001", "번호가 안 채워졌습니다"
    assert filled.email == "김샘플@example.com"
    assert filled.title == "손으로 고친 직함", "앱에서 고친 값을 시트가 덮었습니다"
    assert re.search(r"채운 칸 \d+개", first)

    fill_book(monkeypatch, standing, LIST, owners["a"], "--apply")
    assert "채운 칸 0개" in capsys.readouterr().out, "두 번째 판이 또 얹었습니다"


def test_채우기를_두_번_돌려도_줄이_늘지_않는다(monkeypatch, db, owners, standing):
    """만들기와 같은 약속이다 — 이 파일에서 가장 중요한 성질."""
    fill_book(monkeypatch, standing, LIST, owners["a"], "--apply")
    once = {c.id for c in rows_in(db, LIST)}
    fill_book(monkeypatch, standing, LIST, owners["a"], "--apply")
    assert {c.id for c in rows_in(db, LIST)} == once


def test_채우기는_칸을_만들지_않고_못_세운_칸을_알린다(
        monkeypatch, db, owners, standing, tmp_path, capsys):
    """명함을 찾으러 곁다리 탭을 붙이면 그 탭의 **살림 칸**(`사유`·`전화 여부`)이
    남는 머리글로 읽혀 달 칸으로 선다 — 그 명단에 `9월 딜소개` 옆에 `사유` 가
    나란히 서는 표가 된다.

    줄을 안 만드는 모드가 칸은 만드는 것도 앞뒤가 안 맞는다. 버리지는 않는다 —
    무엇을 못 세웠는지 보여야 사람이 시트를 고치거나 화면에서 칸을 세운다.
    """
    import openpyxl

    from app.services import contact_columns as cc

    side = tmp_path / "side.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "곁다리탭"
    # 이 명단에 없는 머리글들 — 사람이 시트에서 일을 챙기려고 둔 칸이다.
    ws.append(["이름", "투자사명", "휴대폰", "사유", "전화 여부"])
    ws.append(["김샘플", "샘플투자", "010-7000-0001", "딜 소싱 제안", "미진행"])
    wb.save(side)
    wb.close()

    before = [c.label for c in cc.month_columns(db, LIST, today=date(2026, 8, 15))]
    run(monkeypatch, str(side), LIST, owners["a"], "--mode", "fill",
        "--tab", "곁다리탭", "--apply")
    out = capsys.readouterr().out
    db.expire_all()

    assert [c.label for c in cc.month_columns(db, LIST, today=date(2026, 8, 15))] == before, (
        "채우기가 달 칸을 새로 세웠습니다")
    assert "이 명단에 없는 칸" in out and "사유" in out, (
        "못 세운 칸을 조용히 버렸습니다 — 무엇이 빠졌는지 알 수 없습니다")
    # 그러면서 **명함은 채운다.** 못 세운 칸 하나 때문에 그 줄을 통째로
    # 버리면, 곁다리 탭을 붙인 이유가 사라진다.
    assert [c for c in rows_in(db, LIST)
            if c.name == "김샘플"][0].phone == "010-7000-0001"


def test_채우기는_회사명이_다르면_잇지_않는다(
        monkeypatch, db, owners, standing, tmp_path, capsys):
    """이름이 같아도 투자사명이 다르면 **다른 사람일 수 있다.**

    번호가 없는 줄을 잇는 열쇠는 이름+투자사명 하나뿐이라, 이름만으로 이으면
    남의 명함이 남의 줄에 얹힌다.
    """
    other_firm = book(tmp_path,
                      cards=[card_row("김샘플", "010-7000-0044", "다른투자")],
                      deals=[], name="otherfirm.xlsx")
    run(monkeypatch, other_firm, LIST, owners["a"], "--mode", "fill",
        "--tab", "명함탭", "--apply")
    db.expire_all()

    who = [c for c in rows_in(db, LIST) if c.name == "김샘플"][0]
    assert not (who.phone or ""), "회사명이 다른데 번호를 얹었습니다"
    assert "이 명단에 그 줄이 없다" in capsys.readouterr().out


def test_채우기는_못_찾은_값을_빈칸으로_둔다(monkeypatch, db, owners, standing, capsys):
    """**지어내지 않는다.** 시트에 없는 칸은 비어 있는 것이 맞는 상태다."""
    fill_book(monkeypatch, standing, LIST, owners["a"], "--apply")
    db.expire_all()

    who = [c for c in rows_in(db, LIST) if c.name == "김샘플"][0]
    # 명함 탭에 없는 칸들 — 채운 뒤에도 비어 있어야 한다.
    for field in ("office_phone", "office_fax", "address", "card_registered_at"):
        assert not (getattr(who, field) or ""), f"`{field}` 에 없는 값이 들어갔습니다"


def test_채우기는_앱의_다른_명단에_있는_사람을_건드리지_않는다(
        monkeypatch, db, users, owners, standing, tmp_path, capsys):
    """그 번호가 앱의 다른 줄에 이미 있으면, 얹는 순간 **같은 번호가 두 줄**이 된다.

    누가 맡을지는 사람이 정하는 일이라(배정표) 여기서 옮기지 않고 알린다.
    """
    from app.models import SheetOwner, VcContact

    db.add(SheetOwner(label=OTHER, user_id=users["u2"].id))
    db.add(VcContact(user_id=users["u2"].id, source_sheet=OTHER, name="김샘플",
                     firm="샘플투자", phone="010-7000-0001"))
    db.commit()

    fill_book(monkeypatch, standing, LIST, owners["a"], "--apply")
    db.expire_all()

    assert OTHER in (db.query(VcContact)
                     .filter(VcContact.phone == "010-7000-0001").one().source_sheet)
    assert not ([c for c in rows_in(db, LIST) if c.name == "김샘플"][0].phone or "")
    assert "앱의 다른 명단에 있다" in capsys.readouterr().out


def test_채우기_미리보기가_채울_칸_수와_빈칸으로_둘_줄_수를_적는다(
        monkeypatch, db, owners, standing, capsys):
    """**`--apply` 없이는 한 칸도 안 들어간다.** 그리고 무엇이 채워지고 무엇이
    빈칸으로 남는지가 미리 보여야, 넣기 전에 시트를 고칠 수 있다."""
    fill_book(monkeypatch, standing, LIST, owners["a"])
    out = capsys.readouterr().out
    db.expire_all()

    assert re.search(r"채울 칸 \d+개", out)
    assert re.search(r"시트에서 못 찾아 그대로 두는 줄 \d+개", out)
    assert "--apply" in out
    assert not ([c for c in rows_in(db, LIST) if c.name == "김샘플"][0].phone or ""), \
        "미리보기인데 들어갔습니다"


def test_채우기는_명단의_표_배치를_되돌리지_않는다(monkeypatch, db, owners, standing):
    """**명함을 채우려고 부른 명령이 표 모양을 바꾸면 안 된다.**

    투자사 명함 표로 맞춰 둔 명단이 임포트 한 번에 딜공유 표로 돌아가면, 맞춰
    놓은 것이 조용히 풀린다.
    """
    from app.models import SheetOwner
    from app.services import contact_columns as cc

    db.query(SheetOwner).filter(SheetOwner.label == LIST).one().layout = cc.INVESTOR
    db.commit()

    fill_book(monkeypatch, standing, LIST, owners["a"], "--apply")
    db.expire_all()
    assert db.query(SheetOwner).filter(SheetOwner.label == LIST).one().layout \
        == cc.INVESTOR


def test_만들기를_다시_돌려도_표_배치를_되돌리지_않는다(monkeypatch, db, owners, sample):
    """숨김과 같은 이유다 — 화면에서 사람이 정한 값을 임포트 한 번이 되돌리면 안 된다."""
    from app.models import SheetOwner
    from app.services import contact_columns as cc

    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    db.expire_all()
    db.query(SheetOwner).filter(SheetOwner.label == LIST).one().layout = cc.INVESTOR
    db.commit()

    run_book(monkeypatch, sample, LIST, owners["a"], "--apply")
    db.expire_all()
    assert db.query(SheetOwner).filter(SheetOwner.label == LIST).one().layout \
        == cc.INVESTOR


def test_모드를_안_적으면_멈춘다(monkeypatch, db, owners, sample):
    """**기본값을 두지 않는다.** 위험한 쪽(만들기)이 기본이면 채울 자리에
    새 줄이 서고, 안전한 쪽이 기본이면 "왜 아무것도 안 들어갔지" 를 매번 겪는다.
    """
    with pytest.raises(SystemExit):
        run(monkeypatch, sample, LIST, owners["a"], "--tab", "명함탭", "--apply")
    assert rows_in(db, LIST) == []
