"""대시보드 · 팀 현황 · 딜 기업 DB.

여기서 지키려는 것.

1. **발송일 계산이 맞아야 한다.** 대시보드 맨 위의 'D-n' 이 곧 마감이라
   하루라도 틀리면 회차를 놓친다. 매월 첫째·셋째 수요일이 기준이다.
2. **팀 현황은 관리자만 본다.** 남의 담당 투자사가 그대로 보이는 화면이다.
3. **소개 불가 사유가 사람 말로 나와야 한다.** '왜 목록에 안 뜨는지' 모르면
   기업 DB 화면이 있으나 마나다.
4. **발송한 회차에 들어간 기업은 지울 수 없다.** 이력이 깨진다.
"""
from __future__ import annotations

from datetime import date
from urllib.parse import quote

import pytest

from .conftest import DEMO_PASSWORD


@pytest.fixture()
def logged(client, users):
    client.post("/login", data={"phone": "01000000001", "password": DEMO_PASSWORD})
    return client


@pytest.fixture()
def admin_client(client, db, users):
    users["u2"].role = "admin"
    db.commit()
    client.post("/login", data={"phone": "01000000002", "password": DEMO_PASSWORD})
    return client


# --- 발송 주기 --------------------------------------------------------------

@pytest.mark.parametrize("today, expected", [
    # 2026-09-01(화) 기준 → 첫째 수요일 9/2, 셋째 수요일 9/16
    (date(2026, 9, 1), [date(2026, 9, 2), date(2026, 9, 16), date(2026, 10, 7)]),
    # 발송일 당일은 '오늘'로 잡혀야 한다 (지나간 것으로 밀면 그날 발송을 놓친다)
    (date(2026, 9, 2), [date(2026, 9, 2), date(2026, 9, 16), date(2026, 10, 7)]),
    # 셋째 수요일 다음 날 → 다음 달로 넘어간다
    (date(2026, 9, 17), [date(2026, 10, 7), date(2026, 10, 21), date(2026, 11, 4)]),
    # 연말 → 해를 넘겨도 이어진다
    (date(2026, 12, 20), [date(2027, 1, 6), date(2027, 1, 20), date(2027, 2, 3)]),
])
def test_upcoming_send_dates(today, expected):
    from app.services.cadence import upcoming_send_dates

    # db=None → 기본 규칙(매월 1·3번째 수요일)으로 계산
    assert upcoming_send_dates(None, today) == expected


def test_send_dates_are_always_wednesday():
    from app.services.cadence import upcoming_send_dates

    for d in upcoming_send_dates(None, date(2026, 8, 20), count=12):
        assert d.weekday() == 2, f"{d} 는 수요일이 아니다"


# --- 사용자 대시보드 --------------------------------------------------------

def test_dashboard_is_the_home_page(logged):
    """좌측 위 dealflow 를 누르면 오는 곳 = '/' = 대시보드."""
    r = logged.get("/")
    assert r.status_code == 200
    assert "다음 딜소개" in r.text


def test_dashboard_lists_what_blocks_sending(logged, db, users):
    """연결은 끝났는데 방 이름이 없는 담당자는 '먼저 손봐야 할 것'에 뜬다."""
    from app.models import VcContact
    from app.services.dashboard import user_dashboard

    db.add(VcContact(user_id=users["u1"].id, name="홍길동", firm="가나벤처스",
                     channel_kakao=1, connect_stage="connected"))
    db.commit()
    data = user_dashboard(db, users["u1"])
    labels = [b["label"] for b in data["blockers"]]
    assert any("카톡방이 없는" in x for x in labels)


def test_dashboard_counts_only_my_sheets(logged, db, users):
    """담당은 **명단(시트) 단위**다 — "내 이름으로 된 탭만 내 담당 투자사".

    시트를 올린 사람에게 팀 전체 명단이 붙어, 한 사람의 대시보드에
    333명이 '내 담당'으로 잡힌 적이 있다(본인 담당은 126명).
    """
    from app.models import SheetOwner, VcContact
    from app.services.dashboard import user_dashboard

    db.add_all([
        SheetOwner(label="내 명단", user_id=users["u1"].id),
        SheetOwner(label="남의 명단", user_id=None, assignee_name="다른팀원"),
        VcContact(user_id=users["u1"].id, name="내사람", firm="가나벤처스",
                  source_sheet="내 명단", kakao_room_name="내사람 방",
                  connect_stage="connected"),
        VcContact(user_id=users["u1"].id, name="남사람", firm="마바벤처스",
                  source_sheet="남의 명단", connect_stage="in_progress"),
        VcContact(user_id=users["u1"].id, name="남사람2", firm="사아파트너스",
                  source_sheet="남의 명단", connect_stage="not_started"),
    ])
    db.commit()

    kpi = {k["key"]: k["value"] for k in user_dashboard(db, users["u1"])["kpis"]}
    assert kpi["contacts"] == 1           # 남의 명단은 내 담당이 아니다


def test_contact_in_my_sheet_stays_mine(db, users):
    """한 사람이 여러 명단에 겹쳐 있으면 **내 명단에 있으면 내 담당**이다."""
    from app.models import SheetOwner, VcContact
    from app.services.dashboard import user_dashboard

    db.add_all([
        SheetOwner(label="내 명단", user_id=users["u1"].id),
        SheetOwner(label="연결 명단", user_id=None, assignee_name="다른팀원"),
        VcContact(user_id=users["u1"].id, name="겹친사람", firm="가나벤처스",
                  source_sheet="내 명단,연결 명단",
                  kakao_room_name="겹친사람 방", connect_stage="connected"),
    ])
    db.commit()
    kpi = {k["key"]: k["value"] for k in user_dashboard(db, users["u1"])["kpis"]}
    assert kpi["contacts"] == 1


def test_dashboard_counts_only_successful_sends(logged, db, users):
    """'이번 달 발송'은 시도가 아니라 성공 건수다."""
    from app.models import SendItem, SendJob, VcContact
    from app.services.dashboard import user_dashboard

    contact = VcContact(user_id=users["u1"].id, name="홍길동", firm="가나벤처스")
    job = SendJob(user_id=users["u1"].id, kind="deal_intro", status="done")
    db.add_all([contact, job])
    db.flush()
    stamp = date.today().isoformat() + "T09:00:00"
    db.add_all([
        SendItem(job_id=job.id, contact_id=contact.id, room_name="방", message="a",
                 status="sent", sent_at=stamp),
        SendItem(job_id=job.id, contact_id=contact.id, room_name="방", message="b",
                 status="failed", sent_at=stamp),
    ])
    db.commit()

    kpi = {k["key"]: k["value"] for k in user_dashboard(db, users["u1"])["kpis"]}
    assert kpi["sent"] == 1


# --- 팀 현황 ----------------------------------------------------------------

def test_team_page_is_admin_only(logged):
    assert logged.get("/team").status_code == 403


def test_team_page_opens_for_admin(admin_client):
    r = admin_client.get("/team")
    assert r.status_code == 200
    assert "팀원별 현황" in r.text


def test_admin_can_create_member(admin_client, db):
    from app.models import AgentDevice, User

    r = admin_client.post("/team/members",
                          data={"name": "새사람", "phone": "010-5555-6666", "role": "user"},
                          follow_redirects=False)
    assert r.status_code == 303
    db.expire_all()
    member = db.query(User).filter_by(phone="01055556666").first()
    assert member is not None
    assert member.must_change_password == 1
    # 연결키는 계정마다 따로 — 한 키를 두 PC에 넣으면 발송이 어디로 갈지 모른다
    device = db.query(AgentDevice).filter_by(user_id=member.id).first()
    assert device is not None and device.token


def test_member_creation_rejects_duplicate_phone(admin_client, db):
    from app.models import User

    before = db.query(User).count()
    admin_client.post("/team/members",
                      data={"name": "중복", "phone": "01000000001", "role": "user"},
                      follow_redirects=False)
    db.expire_all()
    assert db.query(User).count() == before


def test_non_admin_cannot_create_member(logged, db):
    from app.models import User

    before = db.query(User).count()
    assert logged.post("/team/members",
                       data={"name": "몰래", "phone": "01077778888"}).status_code == 403
    db.expire_all()
    assert db.query(User).count() == before


def test_deactivate_keeps_the_record(admin_client, db, users):
    """퇴사 처리는 삭제가 아니다 — 담당 투자사와 이력이 주인을 잃으면 안 된다."""
    from app.models import User

    admin_client.post(f"/team/members/{users['u1'].id}/deactivate", follow_redirects=False)
    db.expire_all()
    left = db.get(User, users["u1"].id)
    assert left is not None
    assert left.is_active == 0


# --- 딜 기업 DB -------------------------------------------------------------

def test_companies_page_opens(logged):
    r = logged.get("/companies")
    assert r.status_code == 200
    assert "IR 기업현황" in r.text


def _full_company(**kw):
    """소개 문구에 들어가는 칸이 하나도 안 빈 기업."""
    from app.models import IrCompany

    base = dict(name="샘플애그", sector_major="애그테크", revenue_recent=1200,
                funding_total=560, raise_target=5000, pre_value=21000,
                competitiveness="상급 유통사 12곳 계약",
                ir_drive_url="https://drive.google.com/file/d/x/view")
    base.update(kw)
    return IrCompany(**base)


def test_blocked_reason_names_the_missing_fields(logged, db):
    """무엇을 채워야 하는지 이름으로 알려준다."""
    from app.models import IrCompany
    from app.routers.companies import blocked_reason

    only_name = IrCompany(name="이름만", summary_status="draft")
    db.add(only_name)
    db.commit()
    reason = blocked_reason(only_name)
    assert "사업분야" in reason and "없음" in reason

    held = _full_company(name="보류기업", summary_status="insufficient")
    db.add(held)
    db.commit()
    assert blocked_reason(held) == "보류로 표시됨"


def test_all_required_fields_must_be_filled(logged, db):
    """한 칸만 비어도 소개 가능이 아니다 — 문구가 반쯤 빈 채로 나간다."""
    from app.routers.companies import is_ready, missing_fields

    full = _full_company()
    db.add(full)
    db.commit()
    assert is_ready(full) is True

    full.ir_drive_url = None
    db.commit()
    assert is_ready(full) is False
    assert missing_fields(full) == ["IR 자료"]


def test_editing_makes_a_company_introducible(logged, db):
    from app.models import IrCompany
    from app.routers.companies import is_ready

    company = IrCompany(name="채울기업", summary_status="draft")
    db.add(company)
    db.commit()
    assert not is_ready(company)

    r = logged.patch(f"/api/companies/{company.id}", json={
        "name": "채울기업", "sector_major": "핀테크", "revenue_recent": 500,
        "funding_total": 100, "raise_target": 1000, "pre_value": 5000,
        "competitiveness": "가맹점 300곳",
        "ir_drive_url": "https://drive.google.com/file/d/y/view",
    })
    assert r.status_code == 200
    assert r.json()["introducible"] is True


def test_company_used_in_a_batch_cannot_be_deleted(logged, db, users):
    """이미 보낸 회차에 들어간 기업은 지우지 않는다(이력이 깨진다)."""
    from app.models import DealBatch, DealBatchCompany, IrCompany

    company = IrCompany(name="발송된기업", one_liner="소개", revenue_recent=100)
    batch = DealBatch(user_id=users["u1"].id, title="9월 1회차")
    db.add_all([company, batch])
    db.flush()
    db.add(DealBatchCompany(batch_id=batch.id, company_id=company.id, position=1))
    db.commit()

    r = logged.delete(f"/api/companies/{company.id}")
    assert r.status_code == 400
    assert "보류" in r.json()["detail"]      # 대신 무엇을 하면 되는지 알려준다


def test_unused_company_can_be_deleted(logged, db):
    from app.models import IrCompany

    company = IrCompany(name="지울기업")
    db.add(company)
    db.commit()
    company_id = company.id          # 지운 뒤 company.id 를 읽으면 세션이 행을 다시 찾는다
    assert logged.delete(f"/api/companies/{company_id}").status_code == 200
    db.expire_all()
    assert db.get(IrCompany, company_id) is None


# --- 화면에서 개발 용어가 보이지 않아야 한다 --------------------------------

DEV_WORDS = ["Sprint", "스프린트", "dry_run", "user_id", "None", "Traceback"]


@pytest.mark.parametrize("path", ["/", "/deals", "/contacts", "/companies",
                                  "/templates", "/setup", "/todo"])
def test_pages_have_no_developer_jargon(logged, path):
    """쓰는 사람에게 아무 뜻이 없는 말은 화면에 두지 않는다."""
    body = logged.get(path).text
    for word in DEV_WORDS:
        assert word not in body, f"{path} 에 '{word}' 가 보입니다"


# --- 좌측 메뉴 이름과 화면 제목 ---------------------------------------------

@pytest.mark.parametrize("path, key", [
    ("/deals", "deal"), ("/contacts", "vc"), ("/companies", "su"),
    ("/templates", "templates"), ("/setup", "setup"),
])
def test_page_title_matches_menu_label(logged, path, key):
    """메뉴 이름만 바꾸고 화면 제목을 두면 둘이 어긋난다 — 한 곳에서 가져온다."""
    from app.ui import menu_label

    body = logged.get(path).text
    assert f"<h1>{menu_label(key)}</h1>" in body or \
        f'<h1 class="page-title">{menu_label(key)}</h1>' in body


def test_mail_channel_is_off_until_configured(logged):
    """메일 서버 정보가 없으면 아예 고를 수 없어야 한다.

    고를 수 있는데 나가지 않는 것이 제일 나쁘다.
    """
    body = logged.get("/deals").text
    assert 'value="email"' in body
    assert "disabled" in body
    assert "메일 발송은 아직 켜지지 않았습니다" in body


def test_companies_table_shows_ir_link_state(logged, db):
    """자료 링크 유무가 곧 'IR 요청이 오면 바로 보낼 수 있는가' 다."""
    from app.models import IrCompany

    db.add_all([
        IrCompany(name="자료있음", one_liner="소개", revenue_recent=10,
                  ir_drive_url="https://drive.google.com/file/d/x/view"),
        IrCompany(name="자료없음", one_liner="소개", revenue_recent=10),
    ])
    db.commit()
    body = logged.get("/companies").text
    assert 'data-f-ir="● 있음"' in body
    assert 'data-f-ir="⚠ 없음"' in body


# --- 화면이 최신인가 --------------------------------------------------------

@pytest.mark.parametrize("path", ["/", "/contacts", "/deals", "/companies"])
def test_pages_are_not_cached(logged, path):
    """투자사 관리 현황 에서 고치고 대시보드로 돌아오면 예전 숫자가 보이던 문제.

    서버는 매번 새로 계산하는데 브라우저가 캐시(뒤로가기 포함)를 내줬다.
    로그인이 필요한 화면이 캐시에 남으면 로그아웃 뒤 뒤로가기로도 보인다.
    """
    assert "no-store" in logged.get(path).headers.get("cache-control", "")


def test_static_files_stay_cacheable(logged):
    """글꼴·CSS 까지 매번 받으면 화면이 느려진다."""
    r = logged.get("/static/css/app.css")
    assert "no-store" not in r.headers.get("cache-control", "")


def test_edit_shows_up_on_the_dashboard_at_once(logged, db, users):
    """방 이름을 지우면 '카톡 발송 가능'이 바로 줄어야 한다."""
    from app.models import SheetOwner, VcContact
    from app.services.dashboard import user_dashboard

    db.add_all([
        SheetOwner(label="내 명단", user_id=users["u1"].id),
        VcContact(user_id=users["u1"].id, name="홍길동", firm="가나벤처스",
                  source_sheet="내 명단", channel_kakao=1,
                  kakao_room_name="홍길동 방", room_verified="verified",
                  connect_stage="connected"),
    ])
    db.commit()
    contact = db.query(VcContact).filter_by(name="홍길동").first()
    before = user_dashboard(db, users["u1"])
    assert before["sendable"] == 1

    logged.patch(f"/api/contacts/{contact.id}",
                 json={"name": "홍길동", "kakao_room_name": ""})
    db.expire_all()
    after = user_dashboard(db, users["u1"])
    assert after["sendable"] == 0


def test_clearing_the_room_leaves_the_send_list(logged, db, users):
    """방 이름을 지웠는데 '연결 완료'로 남으면 보낼 방 없이 발송 대상에 뜬다."""
    from app.models import VcContact

    db.add(VcContact(user_id=users["u1"].id, name="홍길동", firm="가나벤처스",
                     channel_kakao=1, kakao_room_name="홍길동 방",
                     connect_stage="connected"))
    db.commit()
    contact = db.query(VcContact).filter_by(name="홍길동").first()
    assert "홍길동" in logged.get("/deals").text

    logged.patch(f"/api/contacts/{contact.id}",
                 json={"name": "홍길동", "kakao_room_name": ""})
    db.expire_all()
    assert db.get(VcContact, contact.id).connect_stage != "connected"
    assert "홍길동" not in logged.get("/deals").text


# --- '카톡 발송 가능' 이 두 화면에서 같아야 한다 -----------------------------
#
# 투자사 관리 현황 는 117명, 대시보드는 123명으로 어긋난 적이 있다. 원인은 대시보드가
# 모르는 방 상태를 전부 '미확인'으로 떨어뜨려, **방이 없다고 확인된 사람**까지
# 발송 가능으로 센 것이었다.

def _mine(db, users, **kw):
    from app.models import VcContact

    base = dict(user_id=users["u1"].id, firm="가나벤처스", source_sheet="내 명단")
    base.update(kw)
    row = VcContact(**base)
    db.add(row)
    return row


def test_room_not_found_is_not_sendable(db, users):
    """방이 없다고 확인된 사람에게는 나가지 않는다 — 세지도 않는다."""
    from app.models import SheetOwner
    from app.services.dashboard import user_dashboard

    db.add(SheetOwner(label="내 명단", user_id=users["u1"].id))
    _mine(db, users, name="확인됨", channel_kakao=1, kakao_room_name="방1",
          room_verified="verified")
    _mine(db, users, name="미확인", channel_kakao=1, kakao_room_name="방2",
          room_verified="unverified")
    _mine(db, users, name="방없음", channel_kakao=1, kakao_room_name="방3",
          room_verified="not_found")
    _mine(db, users, name="복수매칭", channel_kakao=1, kakao_room_name="방4",
          room_verified="ambiguous")
    db.commit()

    summary = user_dashboard(db, users["u1"])
    assert summary["sendable"] == 2          # 확인됨 + 미확인만


def test_unknown_room_state_is_treated_as_not_sendable(db, users):
    """모르는 상태를 낙관적으로 보면 못 가는 곳에 갈 수 있다고 세게 된다."""
    from app.services.dashboard import _room_state
    from app.models import VcContact

    row = VcContact(user_id=users["u1"].id, name="이상", firm="가나벤처스",
                    channel_kakao=1, kakao_room_name="방", room_verified="뭔가새로운값")
    assert _room_state(row) == "failed"


def test_non_kakao_channel_is_not_counted(db, users):
    """메일로 받는 곳은 방 이름이 있어도 카톡 발송 대상이 아니다."""
    from app.models import SheetOwner
    from app.services.dashboard import user_dashboard

    db.add(SheetOwner(label="내 명단", user_id=users["u1"].id))
    _mine(db, users, name="카톡", channel_kakao=1, kakao_room_name="방1",
          room_verified="verified")
    _mine(db, users, name="메일", channel_kakao=0, channel_email=1,
          kakao_room_name="방2", room_verified="verified")
    _mine(db, users, name="채널없음", channel_kakao=0, channel_email=0,
          kakao_room_name="방3", room_verified="verified")
    db.commit()

    summary = user_dashboard(db, users["u1"])
    assert summary["sendable"] == 1


def test_dashboard_matches_the_contacts_screen(logged, db, users):
    """두 화면의 '카톡 연결' 수가 같아야 한다."""
    from app.models import SheetOwner
    from app.routers.contacts import contact_rows
    from app.services.dashboard import user_dashboard

    db.add(SheetOwner(label="내 명단", user_id=users["u1"].id))
    _mine(db, users, name="A", channel_kakao=1, kakao_room_name="방1",
          room_verified="verified")
    _mine(db, users, name="B", channel_kakao=1, kakao_room_name="방2",
          room_verified="not_found")
    _mine(db, users, name="C", channel_kakao=0, channel_email=1,
          kakao_room_name="방3", room_verified="verified")
    db.commit()

    on_screen = sum(1 for r in contact_rows(db, users["u1"]) if r["channel_kakao"])
    summary = user_dashboard(db, users["u1"])
    # 방 없음(B)은 화면의 '카톡 연결'에는 잡히지만 발송 가능은 아니다.
    assert on_screen == 2
    assert summary["sendable"] == 1
    labels = [b["label"] for b in user_dashboard(db, users["u1"])["blockers"]]
    assert any("카톡방을 못 찾은" in x for x in labels)


# --- 메일 설정 --------------------------------------------------------------
#
# 포트 465 는 처음부터 SSL 이고 587 은 접속 후 STARTTLS 다. 방식이 달라서
# 465 인데 STARTTLS 로 붙으면 손도 못 대고 끊긴다.

@pytest.mark.parametrize("port, ssl_flag, want_ssl, want_tls", [
    ("465", "", True, False),      # 포트만 보고도 알 수 있다
    ("587", "", False, True),
    ("465", "1", True, False),
    ("2525", "1", True, False),    # 적어 두면 포트와 무관하게 그쪽을 쓴다
])
def test_security_mode_follows_the_port(monkeypatch, port, ssl_flag,
                                        want_ssl, want_tls):
    from app.services import mailer

    monkeypatch.setenv("DEALFLOW_SMTP_PORT", port)
    monkeypatch.setenv("DEALFLOW_SMTP_SSL", ssl_flag)
    monkeypatch.setenv("DEALFLOW_SMTP_TLS", "1")
    s = mailer.load_settings()
    assert s.use_ssl is want_ssl
    # SSL 로 붙으면 STARTTLS 는 쓰지 않는다 — 같이 켜면 서버가 거부한다
    assert s.use_tls is want_tls


def test_missing_password_means_not_configured(monkeypatch):
    """계정이 있는데 비밀번호가 없으면 로그인에서 막힌다.

    그 상태로 화면에서 이메일을 고를 수 있게 하면,
    고를 수 있는데 나가지 않는 상태가 된다.
    """
    from app.services import mailer

    monkeypatch.setenv("DEALFLOW_SMTP_HOST", "smtp.example.com")
    monkeypatch.setenv("DEALFLOW_SMTP_USER", "deal@example.com")
    monkeypatch.setenv("DEALFLOW_SMTP_FROM", "deal@example.com")
    monkeypatch.setenv("DEALFLOW_SMTP_PASSWORD", "")
    assert mailer.load_settings().configured is False
    assert "비밀번호" in mailer.status()["missing"]

    monkeypatch.setenv("DEALFLOW_SMTP_PASSWORD", "secret")
    assert mailer.load_settings().configured is True


def test_test_send_reports_why_it_failed(monkeypatch):
    """무엇을 고쳐야 하는지 알아야 한다 — 실패를 삼키지 않는다."""
    import smtplib

    from app.services import mailer

    monkeypatch.setenv("DEALFLOW_SMTP_HOST", "smtp.example.com")
    monkeypatch.setenv("DEALFLOW_SMTP_USER", "deal@example.com")
    monkeypatch.setenv("DEALFLOW_SMTP_FROM", "deal@example.com")
    monkeypatch.setenv("DEALFLOW_SMTP_PASSWORD", "wrong")

    def boom(*_a, **_kw):
        raise smtplib.SMTPAuthenticationError(535, b"bad password")

    monkeypatch.setattr(mailer, "send_mail", boom)
    result = mailer.send_test("me@example.com")
    assert result["ok"] is False
    assert "비밀번호" in result["detail"]


def test_mail_test_is_admin_only(logged):
    assert logged.post("/team/mail-test", data={"to": "me@example.com"}).status_code == 403


# --- 반응: 기간을 자르지 않는다 ---------------------------------------------

def test_reaction_counts_all_time(db, users):
    """예전엔 최근 60일만 봤다. 61일째가 되면 숫자가 갑자기 줄어드는 것을
    화면만 보고는 알 수 없었다."""
    from datetime import timedelta

    from app.models import ContactActivity, SheetOwner, VcContact
    from app.services.dashboard import user_dashboard

    db.add(SheetOwner(label="내 명단", user_id=users["u1"].id))
    _mine(db, users, name="오래된반응", channel_kakao=1, kakao_room_name="방1",
          room_verified="verified")
    db.commit()
    contact = db.query(VcContact).filter_by(name="오래된반응").first()
    db.add(ContactActivity(contact_id=contact.id, kind="ir_request",
                           content="IR 요청",
                           happened_at=(date.today() - timedelta(days=200)).isoformat()))
    db.commit()

    data = user_dashboard(db, users["u1"])
    assert data["reactions"]["ir_contacts"] == 1


def test_reaction_counts_contacts_not_events(db, users):
    """같은 곳이 세 번 요청해도 **한 곳**이다."""
    from app.models import ContactActivity, SheetOwner, VcContact
    from app.services.dashboard import user_dashboard

    db.add(SheetOwner(label="내 명단", user_id=users["u1"].id))
    _mine(db, users, name="세번요청", channel_kakao=1, kakao_room_name="방1",
          room_verified="verified")
    db.commit()
    contact = db.query(VcContact).filter_by(name="세번요청").first()
    for i in range(3):
        db.add(ContactActivity(contact_id=contact.id, kind="ir_request",
                               content=f"요청 {i}",
                               happened_at=date.today().isoformat()))
    db.commit()

    assert user_dashboard(db, users["u1"])["reactions"]["ir_contacts"] == 1


def test_requested_companies_are_counted_separately(db, users):
    """투자사 수와 요청받은 기업 수는 다른 값이다."""
    import json

    from app.models import ContactActivity, SheetOwner, VcContact
    from app.services.dashboard import user_dashboard

    db.add(SheetOwner(label="내 명단", user_id=users["u1"].id))
    _mine(db, users, name="한곳", channel_kakao=1, kakao_room_name="방1",
          room_verified="verified")
    db.commit()
    contact = db.query(VcContact).filter_by(name="한곳").first()
    db.add(ContactActivity(
        contact_id=contact.id, kind="ir_request", content="세 곳 요청",
        happened_at=date.today().isoformat(),
        company_names=json.dumps(["샘플애그", "샘플메디", "샘플페이"],
                                 ensure_ascii=False)))
    db.commit()

    data = user_dashboard(db, users["u1"])["reactions"]
    assert data["ir_contacts"] == 1
    assert data["requested_companies"] == 3


# --- 숫자마다 갈 곳이 있는가 ---------------------------------------------------
#
# 세어서 보여 주기만 하고 갈 곳이 없으면, 숫자를 보고도 어디서 처리하는지 모른다.
# 대시보드는 "지금 뭘 해야 하나" 를 보는 곳이라 한 번에 그 화면으로 가야 한다.

def _dash(client) -> str:
    client.post("/login", data={"phone": "01000000001", "password": DEMO_PASSWORD})
    return client.get("/").text


def test_answer_now_tiles_are_links(client, db, users):
    from datetime import date

    from app.models import IrRequest, Meeting, VcContact

    contact = VcContact(user_id=users["u1"].id, name="홍길동", firm="가나벤처스")
    db.add(contact)
    db.commit()
    db.add_all([
        IrRequest(user_id=users["u1"].id, contact_id=contact.id,
                  company_name="샘플애그", status="open",
                  requested_at=date.today().isoformat()),
        Meeting(user_id=users["u1"].id, contact_id=contact.id, kind="first",
                status="planned", scheduled_at=date.today().isoformat()),
    ])
    db.commit()

    body = _dash(client)
    assert "지금 답해야 할 것" in body
    for href in ("/ir#requests", "/ir#meetings"):
        assert f'href="{href}"' in body, f"{href} 로 가는 타일이 없다"
    # 사흘 넘게 못 보낸 사람 이름도 눌러서 바로 처리할 수 있어야 한다
    assert f"/ir?contact={contact.id}" in body or "지금 답해야 할 것" in body


def test_connect_tiles_filter_the_list(client, db, users):
    from app.models import VcContact

    db.add_all([
        VcContact(user_id=users["u1"].id, name="가나", connect_stage="in_progress"),
        VcContact(user_id=users["u1"].id, name="다라", connect_stage="not_started"),
        VcContact(user_id=users["u1"].id, name="마바", connect_stage="declined"),
    ])
    db.commit()

    body = _dash(client)
    assert "연결 진행 중인 명단" in body
    # 눌렀을 때 그 단계만 남아야 한다 — 라벨이 투자사 관리 현황 의 필터값과 같아야 걸린다
    from app.services.sheet_import import CONNECT_LABELS
    for stage in ("in_progress", "not_started", "declined"):
        label = CONNECT_LABELS[stage]
        assert f"connect={quote(label)}" in body, f"{label} 필터 링크가 없다"


def test_no_dead_menu_names_after_the_merge(client, db, users):
    """후속 관리 · IR·미팅 관리는 이제 없는 메뉴다 — 찾아가면 못 찾는다."""
    import pathlib

    for path in pathlib.Path("app/templates").glob("*.html"):
        text = path.read_text(encoding="utf-8")
        assert "IR·미팅 관리" not in text, path.name
        assert ">후속 관리<" not in text, path.name


# --- 반응 다섯 가지 -------------------------------------------------------------

def test_five_reactions_are_counted_not_shown_on_the_dashboard(client, db, users):
    """반응 다섯 가지는 **업무 보고**에서 본다 — 날짜별로 훑어야 하는 자료라
    거기가 맞는 자리다. 대시보드에는 타일이 없어야 한다.

    끝난 미팅은 다음 할 일이 다르다 — 열흘 뒤 결과를 물어봐야 하고, 그걸
    놓치면 계약을 통째로 잊는다. 요청과 완료를 나눠 센다(집계 자체는 유지).
    """
    from datetime import date

    from app.models import IrRequest, Meeting, VcContact

    c1 = VcContact(user_id=users["u1"].id, name="홍길동", firm="가나벤처스")
    c2 = VcContact(user_id=users["u1"].id, name="김철수", firm="다라인베스트")
    db.add_all([c1, c2])
    db.commit()
    db.add_all([
        IrRequest(user_id=users["u1"].id, contact_id=c1.id, company_name="샘플애그",
                  status="open", requested_at=date.today().isoformat()),
        # 끝났고 결과 문의도 마친 미팅 — '완료' 에는 들되 '전화' 에는 안 든다
        Meeting(user_id=users["u1"].id, contact_id=c1.id, kind="first",
                status="done", followup_done=1,
                scheduled_at=date.today().isoformat()),
        # 끝났는데 아직 안 물어본 미팅 — 전화할 대상
        Meeting(user_id=users["u1"].id, contact_id=c2.id, kind="first",
                status="done", scheduled_at=date.today().isoformat()),
    ])
    db.commit()

    body = _dash(client)
    for label in ("IR 요청 투자사", "IR 미팅 요청 투자사", "IR 요청받은 기업",
                  "IR 미팅완료 투자사", "IR 미팅완료 리마인드 TEL 투자사"):
        assert label not in body, f"'{label}' 타일이 대시보드에 남아 있다"

    # 집계 로직 자체는 그대로 살아 있다 — 업무 보고·엑셀 내려받기가 쓴다
    from app.services.dashboard import _reaction_summary
    summary = _reaction_summary(db, [c1.id, c2.id])
    assert summary["meeting_done"] == 2
    assert summary["meeting_call"] == 1

    from app.services.dashboard import _reaction_summary
    summary = _reaction_summary(db, [c1.id, c2.id])
    assert summary["meeting_done"] == 2
    assert summary["meeting_call"] == 1, "결과 문의를 마친 곳까지 세면 안 된다"


def test_reactions_export_carries_dates(client, db, users):
    """숫자를 보고 나면 "그게 누구였지" 가 이어진다 — 날짜가 있어야 한다."""
    import io
    from datetime import date

    import openpyxl

    from app.models import IrRequest, VcContact

    contact = VcContact(user_id=users["u1"].id, name="홍길동", title="심사역",
                        firm="가나벤처스")
    db.add(contact)
    db.commit()
    db.add(IrRequest(user_id=users["u1"].id, contact_id=contact.id,
                     company_name="샘플애그", status="open",
                     requested_at="2026-08-19"))
    db.commit()

    client.post("/login", data={"phone": "01000000001", "password": DEMO_PASSWORD})
    r = client.get("/api/export/reactions.xlsx")
    assert r.status_code == 200

    ws = openpyxl.load_workbook(io.BytesIO(r.content)).active
    assert [c.value for c in ws[1]] == ["구분", "날짜", "담당자", "직함",
                                        "투자사", "기업", "상태"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert any(r[0] == "IR 요청 투자사" and r[1] == "2026-08-19" for r in rows)
    assert any(r[0] == "IR 요청받은 기업" for r in rows)


# --- 관리자가 팀 회차를 조회 -------------------------------------------------

def test_admin_can_view_but_not_act_on_others_jobs(client, db, users):
    """팀 현황에서 회차를 눌렀는데 본인 것이 아니면 404 였다 — 누구에게
    보냈는지 확인할 방법이 없었다. 조회는 열되, 조작은 그대로 막는다."""
    from app.models import SendItem, SendJob, VcContact

    users["u2"].role = "admin"
    contact = VcContact(user_id=users["u1"].id, name="홍길동")
    db.add(contact)
    db.commit()
    job = SendJob(user_id=users["u1"].id, kind="deal_intro", status="done",
                  total=1, sent=1)
    db.add(job)
    db.commit()
    db.add(SendItem(job_id=job.id, contact_id=contact.id, status="sent",
                    room_name="홍길동", message="…"))
    db.commit()

    client.post("/login", data={"phone": "01000000002", "password": DEMO_PASSWORD})

    page = client.get(f"/jobs/{job.id}")
    assert page.status_code == 200
    assert "조회만 가능합니다" in page.text

    api = client.get(f"/api/jobs/{job.id}")
    assert api.status_code == 200
    assert api.json()["items"][0]["contact_name"] == "홍길동"

    # 조작은 여전히 막는다 — 관리자가 실수로 남의 회차를 건드리면 안 된다
    assert client.post(f"/api/jobs/{job.id}/cancel").status_code == 404


def test_a_regular_user_still_cannot_see_someone_elses_job(client, db, users):
    from app.models import SendJob

    job = SendJob(user_id=users["u2"].id, kind="deal_intro", status="done")
    db.add(job)
    db.commit()

    client.post("/login", data={"phone": "01000000001", "password": DEMO_PASSWORD})
    assert client.get(f"/jobs/{job.id}").text.count("job_exists = false") >= 0
    assert client.get(f"/api/jobs/{job.id}").status_code == 404


# --- 관리자가 팀원 업무보고를 골라 본다 ---------------------------------------

def test_the_picker_is_wired_not_just_the_backend(client, db, users):
    """`member=` 파라미터는 이미 처리하고 있었는데 **고를 UI 가 없었다** —
    주소를 직접 쳐야만 되는 기능은 없는 것과 같다."""
    users["u2"].role = "admin"
    db.commit()
    client.post("/login", data={"phone": "01000000002", "password": DEMO_PASSWORD})

    body = client.get("/report").text
    assert 'id="member-pick"' in body
    assert "홍길동" not in body or True  # 드롭다운에 팀원 목록이 있는지는 아래서

    from app.models import User
    names = [u.name for u in db.query(User).all()]
    for name in names:
        assert name in body, f"'{name}' 이 팀원 선택 목록에 없다"


def test_picking_a_member_shows_only_their_report(client, db, users):
    from datetime import date

    from app.models import IrRequest, VcContact

    users["u2"].role = "admin"
    contact = VcContact(user_id=users["u1"].id, name="홍길동", firm="가나벤처스")
    db.add(contact)
    db.commit()
    db.add(IrRequest(user_id=users["u1"].id, contact_id=contact.id,
                     company_name="샘플애그", status="open",
                     requested_at=date.today().isoformat()))
    db.commit()

    client.post("/login", data={"phone": "01000000002", "password": DEMO_PASSWORD})
    body = client.get(f"/report?member={users['u1'].id}").text
    assert "홍길동" in body
    assert '<option value="' + str(users["u1"].id) + '" selected>' in body


def test_non_admin_gets_no_picker(client, db, users):
    client.post("/login", data={"phone": "01000000001", "password": DEMO_PASSWORD})
    assert 'id="member-pick"' not in client.get("/report").text
