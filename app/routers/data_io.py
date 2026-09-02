"""투자사 관리 현황 업로드 · 표 엑셀 내려받기.

**업로드**
팀원마다 자기 '투자사 관리 현황' 시트를 따로 관리한다. 그동안은 내가 구글 시트를
직접 읽어 임포트했지만, 팀원이 자기 시트를 스스로 올릴 수 있어야 운영이 돈다.
올린 파일은 기존 임포트 파서(sheet_import.parse_sheet_a)가 그대로 읽는다 —
구글 시트든 내려받은 엑셀이든 같은 표이기 때문이다.

두 단계로 나눈다. **미리보기(dry_run)** 로 무엇이 생기고 무엇이 바뀌는지 먼저 보고,
그 다음에 반영한다. 담당자 명단은 발송 대상이라 잘못 덮으면 그대로 오발송이 된다.

업로드한 사람의 담당분으로 붙는다. 시트에 '담당자' 컬럼이 있으면 그 이름으로 계정을
찾고, 못 찾으면 **버리지 않고** 올린 사람에게 붙인 뒤 리포트에 남긴다.

**내려받기**
화면에서 보는 표를 그대로 엑셀로 준다. 필터·정렬은 엑셀에서 하도록 머리행을 고정하고
자동 필터를 걸어 둔다.

**업무 보고 리포트**(`/api/export/report.xlsx`)만 모양이 다르다. 그건 표 하나가
아니라 화면 한 장을 옮긴 것이고, 카톡으로 쓰던 업무보고를 대신하는 문서라
제목·섹션·인쇄 설정이 붙는다. 자세한 이유는 파일 끝 그 자리에 적어 두었다.
"""
from __future__ import annotations

import io
from datetime import date
from typing import List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import Response
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..db import get_db
from ..deps import get_current_user, may_manage_team_contacts
from ..models import IrCompany, SendItem, SendJob, User, VcContact
from ..services import (report as report_svc, sheet_import, sheet_owner,
                        spreadsheet as sp)
from .contacts import contact_rows

# 경로를 /api/import, /api/export 로 따로 둔다.
# /api/contacts/export.xlsx 로 두면 기존 /api/contacts/{contact_id} 가 먼저 잡아
# "export.xlsx 는 정수가 아니다" 로 422 가 난다 — 등록 순서로 푸는 대신 경로를 가른다.
router = APIRouter(tags=["data"])


# --- 업로드 -----------------------------------------------------------------

def _read_upload(file: UploadFile, sheet: Optional[str]) -> tuple:
    data = file.file.read()
    # 내려받은 표를 그대로 다시 올리면 활동 이력이 뻥튀기된다 — 내보내기의
    # 'IR 요청(누적)' 류 컬럼을 임포트 파서가 월별 활동으로 읽기 때문이다.
    # (127명짜리 내보내기를 되올려 활동 635건이 새로 잡히는 것을 확인했다)
    if sp.is_export_file(file.filename or "", data):
        raise HTTPException(
            status_code=400,
            detail="이 파일은 dealflow 에서 내려받은 표입니다. "
                   "업로드에는 원본 '투자사 관리 현황' 시트를 올려주세요.",
        )
    try:
        rows = sp.read_rows(file.filename or "", data, sheet or None)
        names = sp.sheet_names(file.filename or "", data)
    except sp.SpreadsheetError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    if not rows:
        raise HTTPException(status_code=400, detail="시트에 내용이 없습니다.")
    return rows, names


@router.post("/api/import/contacts")
def import_contacts(
    file: UploadFile = File(...),
    sheet: str = Form(""),
    year: int = Form(0),
    dry_run: bool = Form(True),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """투자사 관리 현황 업로드. 기본은 미리보기 — 반영하려면 dry_run=false."""
    rows, names = _read_upload(file, sheet)

    try:
        parsed = sheet_import.parse_sheet_a(rows, year or date.today().year)
    except ValueError as exc:
        # 헤더를 못 찾는 경우가 대부분이다 — 어느 시트를 골라야 하는지 알려준다.
        detail = str(exc)
        if names:
            detail += f" · 이 파일의 시트: {', '.join(names)}"
        raise HTTPException(status_code=400, detail=detail)

    report = sheet_import.apply_sheet_a(
        db, parsed, user_id=user.id, dry_run=dry_run,
        source_label=(sheet or file.filename or "업로드"),
    )
    if dry_run:
        db.rollback()   # 미리보기가 DB 를 건드리고 끝나면 안 된다

    return {
        "dry_run": dry_run,
        "sheets": names,
        "sheet_used": sheet or (names[0] if names else ""),
        "header_row": parsed.header_row,
        "parsed_contacts": len(parsed.contacts),
        "created": report.created,
        "updated": report.updated,
        "activities_created": report.activities_created,
        "activities_existing": report.activities_existing,
        "notes": report.notes,
        "skipped": [{"row": s.row_no, "reason": s.reason, "preview": s.preview}
                    for s in report.skipped[:50]],
        "skipped_total": len(report.skipped),
        "summary": report.as_text("업로드"),
    }


@router.post("/api/import/contacts/sheets")
def list_upload_sheets(file: UploadFile = File(...)):
    """파일을 올리기 전에 어떤 시트가 들어 있는지 보여준다."""
    data = file.file.read()
    try:
        return {"sheets": sp.sheet_names(file.filename or "", data)}
    except sp.SpreadsheetError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


# --- 내려받기 ---------------------------------------------------------------

def _xlsx(filename: str, sheet_title: str, headers, rows) -> Response:
    try:
        content = sp.write_xlsx(sheet_title, headers, rows)
    except sp.SpreadsheetError as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    return Response(content=content, media_type=sp.XLSX_MEDIA_TYPE,
                    headers=sp.content_disposition(filename))


# 시트에 있는 값은 **하나도 빠뜨리지 않는다.** 엑셀로 내보낸 뒤 원본 시트와
# 나란히 놓고 대조하는 일이 잦은데, 빠진 칸이 있으면 그때 알게 된다.
CONTACT_HEADERS = [
    "담당 팀원", "담당자", "연결 단계", "그룹", "이름", "직함", "투자사", "부서",
    "채널", "카톡방", "방 확인", "초대", "관심도",
    "라운드 규모", "선호 단계", "선호 분야",
    "휴대폰", "전자 메일 주소", "근무처 전화", "근무처 팩스", "근무지 주소", "명함 등록일",
    "마지막 딜소개", "회차 메모", "IR 요청(최근)", "미팅(최근)",
    "IR 요청(누적)", "미팅(누적)", "상태", "메모",
]


# 업로드용 샘플 양식.
#
# 실제 명단 시트와 **같은 모양**이어야 한다: 머리글이 2행이고, 1행에는 월
# 라벨이 온다(`8월` 이 딜소개/IR/미팅 3열 위에 걸친다). 파서가 위치가 아니라
# 머리글 이름과 그 위아래 문맥으로 컬럼을 찾기 때문에, 이름이 다르면 값이
# 통째로 버려진다 — 샘플이 실제로 통과하는 양식이어야 하는 이유다.
#
# 채워 넣은 사람은 전부 가짜다.
SAMPLE_MONTH_ROW = [
    "", "", "", "", "", "", "", "", "", "", "", "", "",
    "8월", "", "",
]
SAMPLE_HEADER_ROW = [
    "NO", "그룹", "이름", "직책", "투자사명", "부서", "담당자",
    "휴대폰", "전자 메일 주소", "근무처 전화", "근무지 주소",
    "선호 투자분야", "라운드 규모",
    "1차 딜소개", "IR 요청", "미팅",
]
SAMPLE_ROWS = [
    [1, "A그룹", "홍길동", "심사역", "가나벤처스", "투자1본부", "김담당",
     "010-0000-0001", "hong@example.com", "02-000-0001", "서울 강남구 000",
     "AI, 헬스케어", "10~30억",
     "8/13(목) 샘플가, 샘플나", "8/19(수) 2번", ""],
    [2, "A그룹", "김서연", "팀장", "다라인베스트", "", "김담당",
     "010-0000-0002", "kim@example.com", "", "",
     "소재부품, 로봇", "30~50억",
     "8/13(목) 핵심 딜 8개사", "", "8/26(수) 20분"],
    [3, "", "박지훈", "대표", "마바캐피탈", "", "이담당",
     "010-0000-0003", "", "", "",
     "", "", "", "", ""],
]
# 칸이 좁으면 머리글이 잘려서 무엇을 적는 칸인지 안 보인다.
SAMPLE_WIDTHS = [5, 9, 10, 9, 18, 12, 9, 15, 22, 15, 22, 18, 13, 24, 18, 18]


@router.get("/api/sample/contacts.xlsx")
def sample_contacts(user: User = Depends(get_current_user)):
    """투자사 관리 현황 업로드용 **빈 양식**.

    내보내기(`/api/export/contacts.xlsx`)와 다른 파일이다. 내보낸 파일에는
    되올리기 방지 표식이 붙어 있어 그대로 올리면 막힌다 — 새로 명단을 만드는
    사람에게는 올릴 수 있는 양식이 따로 있어야 한다.
    """
    try:
        content = sp.write_plain_xlsx(
            "투자사 관리 현황",
            [SAMPLE_MONTH_ROW, SAMPLE_HEADER_ROW, *SAMPLE_ROWS],
            head_row=2, widths=SAMPLE_WIDTHS)
    except sp.SpreadsheetError as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    return Response(content=content, media_type=sp.XLSX_MEDIA_TYPE,
                    headers=sp.content_disposition("투자사 관리 현황_업로드양식.xlsx"))


def _contact_row(r: dict) -> list:
    """투자사 한 줄 → 엑셀 한 줄. `CONTACT_HEADERS` 와 **같은 차례**다.

    한 곳에 둔 이유는 이 줄을 쓰는 자리가 둘이기 때문이다 — `전체` 에서 받는
    파일과, 표가 템플릿에 적혀 있는 배치(`INVESTOR_LAYOUT`)의 명단별 파일.
    두 곳에 나눠 적으면 칸이 하나 늘 때 한쪽만 늘고, 그 파일은 **머리글과 값이
    한 칸씩 밀린 채로** 나간다 — 열어 봐도 티가 잘 안 난다.
    """
    marks = []
    if r["channel_kakao"]:
        marks.append("카톡")
    if r["channel_email"]:
        marks.append("메일")
    return [r["owner"], r["assignee"], r["connect_label"], r["group_name"],
            r["name"], r["title"], r["firm"], r["department"], "/".join(marks),
            r["room_name"], r["room_label"], r["invited_status"], r["interest_level"],
            r["round_size"], ", ".join(r["stages"]), ", ".join(r["sectors"]),
            r["phone"], r["email"], r["office_phone"], r["office_fax"],
            r["address"], r["card_registered_at"],
            r["last_deal"] or "", r["last_deal_note"],
            r["ir_recent"], r["meet_recent"], r["ir_total"], r["meet_total"],
            r["status_label"], r["memo"]]


def _export_sheet(db: Session, user: User, label: str) -> Response:
    """명단 하나를 **그 명단이 화면에서 쓰는 칸 그대로** 내보낸다.

    명단마다 표가 다르다(`SheetOwner.layout`). 아래 투자사 표 하나로만 내보내면
    스타트업 리마인드 명단을 받았을 때 `부서`·`근무처 팩스` 같은 빈 칸이 스무
    개 오고, 정작 매달 채우는 칸(`8월 리마인드 문자`·`계약여부`)은 **한 칸도
    안 나온다.** 화면에서 보던 것과 다른 파일이 오면 열어 보고 나서야 안다.

    칸 목록은 화면과 **같은 함수**에서 가져온다(`panel_columns`) — 여기에 이름을
    다시 적으면 칸이 하나 늘 때 엑셀에서만 조용히 빠진다. 표에서 뺀 칸까지
    담는 이유는, 엑셀은 가로로 미는 부담이 없어 시트와 대조하는 자리이기
    때문이다(투자사 표도 화면에 없는 칸까지 내보낸다).

    달마다 늘어나는 칸은 **접지 않고 전부** 넣는다 — 화면에서 접는 것은 가로로
    밀리기 때문이고, 파일에는 그 이유가 없다. 지난달 기록이 빠진 파일을 시트와
    대조하면 지워진 것으로 읽힌다.
    """
    from ..services import contact_columns as cc

    layout = cc.layout_of(sheet_owner.layout_of(db, label))
    columns = [c for c in cc.panel_columns(layout, cc.month_columns(db, label))
               if c.source in ("field", "note")]
    # **화면과 같은 줄**이다. 감춘 줄은 화면에서 빠져 있으므로 여기서도 뺀다 —
    # 세어 보고 목록에서 찾을 수 없는 줄이 파일에만 있으면 수가 어긋난다.
    rows = [r for r in contact_rows(db, user,
                                    team_wide=may_manage_team_contacts(user),
                                    include_hidden=True)
            if label in r["sheets"] and not r["is_hidden"]]

    # **표가 템플릿에 적혀 있는 배치는 `head` 가 비어 있다.**
    #
    # `INVESTOR_LAYOUT` 이 그렇다 — 그 표는 `contacts.html` 에 그대로 적혀 있고
    # 여기에는 월별 칸만 남는다(그 배치 주석 참고). 그대로 내보내면 `담당 팀원`
    # 과 달 칸만 든 파일이 나간다. **이름도 투자사명도 연락처도 없다.** 받은
    # 사람은 누구의 무엇인지 알 수 없고, 시트와 대조할 수도 없다.
    #
    # 그래서 이런 배치에서는 **투자사 표 한 장(`CONTACT_HEADERS`)을 앞에 세우고**
    # 그 명단의 달 칸을 뒤에 잇는다. `전체` 에서 받는 파일과 같은 칸이라 나란히
    # 놓고 볼 수 있고, 달 칸은 그 명단 것만 붙으므로 시트와도 대조된다.
    if not layout.head and not layout.tail:
        months = [c for c in columns if c.source == "note"]
        body = [
            _contact_row(r) + [r["notes"].get(c.key, "") for c in months]
            for r in rows
        ]
        today = date.today().isoformat()
        return _xlsx(f"{label}_{today}.xlsx", layout.label,
                     CONTACT_HEADERS + [c.label for c in months], body)

    body = [
        [r["owner"]] + [(r["notes"].get(c.key, "") if c.source == "note"
                         else r.get(c.key, "")) for c in columns]
        for r in rows
    ]
    today = date.today().isoformat()
    return _xlsx(f"{label}_{today}.xlsx", layout.label,
                 ["담당 팀원"] + [c.label for c in columns], body)


@router.get("/api/export/contacts.xlsx")
def export_contacts(sheet: str = "", db: Session = Depends(get_db),
                    user: User = Depends(get_current_user)):
    """표 → 엑셀. 화면과 같은 순서·같은 값.

    `sheet` 를 주면 **그 명단이 화면에서 쓰는 칸으로** 나간다(위 `_export_sheet`).
    안 주면 지금까지처럼 투자사 표 한 장이다 — 투자사 관리 현황의 `전체` 에서
    누르는 그 파일이라 모양이 바뀌면 안 된다.

    **없는 명단 이름은 받지 않는다.** 아무 글자나 받으면 빈 파일이 내려가는데,
    받는 쪽은 그 명단에 줄이 없는 것으로 읽는다.
    """
    if (sheet or "").strip():
        name = (sheet or "").strip()
        known = {t["key"] for t in sheet_owner.sheet_rows(
            db, sheet_owner.managed(db, user,
                                    team_wide=may_manage_team_contacts(user),
                                    include_hidden=True))}
        if name not in known:
            raise HTTPException(status_code=404, detail="없는 명단입니다")
        return _export_sheet(db, user, name)

    rows = [
        _contact_row(r)
        # 관리자는 팀 전체를 내려받는다 — 화면과 같은 범위여야 헷갈리지 않는다.
        # **그래서 판정도 화면과 같은 것을 읽는다**(`may_manage_team_contacts`).
        # 여기 `role == "admin"` 을 따로 적어 두면 범위가 갈릴 자리가 하나 더 는다.
        for r in contact_rows(db, user, team_wide=may_manage_team_contacts(user))
    ]
    today = date.today().isoformat()
    return _xlsx(f"내 투자사_{today}.xlsx", "내 투자사", CONTACT_HEADERS, rows)


def _eok(value):
    """저장값(백만원) → 억. 엑셀에서 계산할 수 있게 **숫자로** 둔다."""
    return None if value is None else round(value / 100, 1)


COMPANY_HEADERS = [
    "기업명", "분야(대)", "분야(소)", "기업구분", "한줄소개", "소개가능",
    # 화면과 같은 단위로 내보낸다. 표는 억인데 엑셀만 백만이면, 두 개를 나란히
    # 놓고 보는 사람에게는 같은 값이 100배 차이로 보인다.
    "최근매출(억)", "누적투자(억)", "희망투자(억)", "Pre Value(억)",
    # `계약서 수신됨` 은 **화면 이름 그대로** 적는다. 내려받은 파일을 표와
    # 나란히 놓고 대조하는 자리라, 여기서만 줄여 부르면 어느 칸인지 매번
    # 되짚어야 한다(옆의 `계약`·`계약월` 은 예전부터 쓰던 줄임말이라 그대로
    # 둔다 — 이름을 바꾸면 이 파일을 받아 쓰던 수식이 어긋난다).
    # 자리도 화면과 같이 `계약` 바로 뒤다.
    "경쟁력", "계약", "계약서 수신됨", "계약월", "탑딜", "투자현황", "요약상태",
    "IR 링크",
]


@router.get("/api/export/companies.xlsx")
def export_companies(db: Session = Depends(get_db),
                     user: User = Depends(get_current_user)):
    """IR 기업현황 → 엑셀. 소개 가능 여부를 함께 내보내 어디를 채워야 할지 보이게 한다."""
    companies = db.execute(select(IrCompany).order_by(IrCompany.name)).scalars().all()
    rows = [
        [c.name, c.sector_major or "", c.sector_minor or "", c.series or "",
         c.one_liner or "", "O" if c.introducible else "",
         _eok(c.revenue_recent), _eok(c.funding_total),
         _eok(c.raise_target), _eok(c.pre_value),
         c.competitiveness or "", c.contract_status or "",
         # 아직 안 정한 기업은 **빈 칸**이다 — `X` 로 채우면 엑셀에서 세는
         # 순간 "확인했는데 안 왔다" 가 그 숫자에 들어간다.
         c.contract_received or "", c.contract_month or "",
         "★" if c.is_top_deal else "", c.funding_status or "",
         c.summary_status or "", c.ir_drive_url or ""]
        for c in companies
    ]
    today = date.today().isoformat()
    # 파일 이름과 시트 이름은 **원본 구글 시트의 탭 이름 그대로**다. 좌측 메뉴는
    # `IR 기업 현황` 으로 띄어 쓰지만 여기는 따라가지 않는다 — 내려받은 파일을
    # 그 시트와 나란히 놓고 대조하는 자리라, 이름이 갈리면 어느 탭에 맞춰 볼지
    # 매번 따져야 한다(가져오기도 시트 쪽 이름을 그대로 찾는다:
    # `scripts/import_company_sheets.py` 의 `find_sheet`).
    return _xlsx(f"IR 기업현황_{today}.xlsx", "IR 기업현황", COMPANY_HEADERS, rows)


from ..services.pipeline import REQUEST_STATUS  # noqa: E402

REACTION_HEADERS = ["구분", "날짜", "담당자", "직함", "투자사", "기업", "상태"]


@router.get("/api/export/reactions.xlsx")
def export_reactions(db: Session = Depends(get_db),
                     user: User = Depends(get_current_user)):
    """대시보드의 반응 다섯 가지를 **날짜와 함께** 한 장으로.

    화면은 숫자만 보여준다. 숫자를 보고 나면 "그게 누구였지" 가 이어지는데,
    그때마다 다섯 화면을 돌아다녀야 했다.
    """
    from ..models import IrRequest, Meeting, VcContact
    from ..services.pipeline import MEETING_KINDS, OUTCOMES

    contacts = {
        c.id: c for c in db.execute(
            select(VcContact).where(VcContact.user_id == user.id)).scalars().all()
    }

    def who(contact_id):
        c = contacts.get(contact_id)
        return [c.name, c.title or "", c.firm or ""] if c else ["-", "", ""]

    rows = []
    for r in db.execute(select(IrRequest).where(
            IrRequest.user_id == user.id).order_by(IrRequest.requested_at)).scalars():
        rows.append(["IR 요청 투자사", r.requested_at or "", *who(r.contact_id),
                     r.company_name or "", REQUEST_STATUS.get(r.status, r.status)])
        rows.append(["IR 요청받은 기업", r.requested_at or "", *who(r.contact_id),
                     r.company_name or "", REQUEST_STATUS.get(r.status, r.status)])

    for m in db.execute(select(Meeting).where(
            Meeting.user_id == user.id).order_by(Meeting.scheduled_at)).scalars():
        label = ("IR 미팅완료 투자사" if m.status == "done"
                 else "IR 미팅 요청 투자사")
        rows.append([label, m.scheduled_at or "", *who(m.contact_id),
                     m.company_name or "", MEETING_KINDS.get(m.kind, m.kind)])
        # 끝난 미팅 중 아직 결과를 안 물어본 곳 — 전화할 대상이다.
        if m.status == "done" and not m.followup_done:
            rows.append(["IR 미팅완료 리마인드 TEL 투자사",
                         m.followup_due or m.scheduled_at or "", *who(m.contact_id),
                         m.company_name or "",
                         OUTCOMES.get(m.outcome or "", "결과 미정")])

    today = date.today().isoformat()
    return _xlsx(f"반응 현황_{today}.xlsx", "반응 현황", REACTION_HEADERS, rows)


JOB_HEADERS = ["담당자", "직함", "투자사", "카톡방", "상태", "발송시각", "실패사유", "문구"]


@router.get("/api/export/jobs/{job_id}.xlsx")
def export_job(job_id: int, db: Session = Depends(get_db),
               user: User = Depends(get_current_user)):
    """발송 회차 결과 → 엑셀. 어디로 무엇이 나갔는지 그대로 남긴다."""
    job = db.get(SendJob, job_id)
    if job is None or job.user_id != user.id:
        raise HTTPException(status_code=404, detail="발송 회차를 찾을 수 없습니다")

    items = db.execute(
        select(SendItem).where(SendItem.job_id == job.id).order_by(SendItem.id)
    ).scalars().all()
    contacts = {
        c.id: c for c in db.execute(
            select(VcContact).where(
                VcContact.id.in_([i.contact_id for i in items] or [0]))
        ).scalars().all()
    }
    rows = []
    for item in items:
        c = contacts.get(item.contact_id)
        rows.append([
            c.name if c else "", (c.title or "") if c else "", (c.firm or "") if c else "",
            item.room_name or "", item.status, item.sent_at or "",
            item.error or "", item.message or "",
        ])
    return _xlsx(f"발송결과_{job.id}.xlsx", f"발송 {job.id}", JOB_HEADERS, rows)


# --- 업무 보고 → 리포트 엑셀 -------------------------------------------------
#
# 업무 보고 화면(`/report`)을 **카톡 업무보고 대신** 쓴다. 그러려면 화면에서 읽은
# 것을 손으로 옮겨 적지 않고 파일 하나로 넘길 수 있어야 한다. 그래서 이 파일은
# 화면을 위에서 아래로 그대로 옮긴다 — 요약 · 발송 · 미팅 · 반응.
#
# **시트를 셋으로 나눈 이유.** 엑셀은 칸 폭이 시트 단위다. 발송(회차명 한 줄)과
# 미팅(후기 한 문단)과 반응(투자사·기업)을 한 장에 쌓으면 어느 표도 제 폭을 갖지
# 못해 셋 다 읽기 어려워진다. 화면 순서대로 시트를 세우면 파일을 열었을 때 첫
# 장이 화면 맨 위(요약·발송)이고, 표마다 제 폭을 갖는다.
#
# **꾸미기는 읽는 품을 줄이는 선까지만.** 제목 · 섹션 띠 · 머리글 서식 · 칸 폭 ·
# 안 나간 건 빨강 · 인쇄 설정까지다. 숫자를 읽는 문서라 그 이상은 방해가 된다.
# 색은 세 가지뿐이고 뜻이 화면과 같다 — 중단은 빨강, 나머지 미완은 노랑.

_R_HEAD_FILL = "EEF2F7"    # 표 머리글. 다른 내려받기(spreadsheet.write_xlsx)와 같은 색
_R_BAND_FILL = "DCE6F1"    # 섹션 띠 (화면의 panel-title)
_R_GROUP_FILL = "F5F7FA"   # 그룹 머리 (화면의 bucket-head)
_R_BAD_FILL = "FDECEC"     # 중단된 회차 줄
_R_BAD = "B42318"          # 중단 · 날짜 지남
_R_WARN = "B54708"         # 안 나간 건이 남음
_R_MUTED = "667085"
_R_LINE = "D8DEE9"


def _report_width(text: str) -> int:
    """엑셀 칸 폭 단위로 잰 글자 너비. 한글은 두 배다.

    (`spreadsheet._display_width` 와 같은 셈이다. 그쪽은 그 모듈 안에서만
    쓰라고 밑줄을 달아 둔 함수라 여기서는 끌어다 쓰지 않는다.)
    """
    return sum(2 if ord(ch) > 0x2000 else 1 for ch in str(text))


class _ReportSheet:
    """보고 시트를 위에서 아래로 쌓는 붓.

    행 번호를 손으로 세지 않게 한다 — 섹션을 하나 끼워 넣을 때마다 아래 모든
    좌표를 고쳐야 하면, 고치다 빠뜨린 자리가 그대로 어긋난 표가 된다.
    """

    def __init__(self, ws, widths, cols):
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.properties import PageSetupProperties

        self.ws = ws
        self.cols = cols                 # 이 시트의 표가 쓰는 칸 수
        self.at = 1
        self._letter = get_column_letter
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        # 눈금선을 끈다. 표에는 테두리를 직접 그리므로, 눈금선까지 있으면 어디까지가
        # 표인지 안 보인다 — 인쇄해서 보는 문서다.
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0     # 세로는 몇 장이 되든 그대로
        ws.page_margins.left = ws.page_margins.right = 0.4
        ws.page_margins.top = ws.page_margins.bottom = 0.5

    # ── 낱개 ──────────────────────────────────────────────────────────────
    def _merge(self, row):
        if self.cols > 1:
            self.ws.merge_cells(f"A{row}:{self._letter(self.cols)}{row}")

    def blank(self, height=6):
        self.ws.row_dimensions[self.at].height = height
        self.at += 1

    def title(self, text, sub=""):
        from openpyxl.styles import Alignment, Font

        cell = self.ws.cell(row=self.at, column=1, value=text)
        cell.font = Font(bold=True, size=15)
        cell.alignment = Alignment(vertical="center")
        self._merge(self.at)
        self.ws.row_dimensions[self.at].height = 24
        self.at += 1
        if sub:
            cell = self.ws.cell(row=self.at, column=1, value=sub)
            cell.font = Font(size=10, color=_R_MUTED)
            self._merge(self.at)
            self.at += 1

    def band(self, text):
        """섹션 띠 — 화면의 panel-title 자리."""
        from openpyxl.styles import Alignment, Font, PatternFill

        cell = self.ws.cell(row=self.at, column=1, value=text)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill("solid", fgColor=_R_BAND_FILL)
        cell.alignment = Alignment(vertical="center")
        self._merge(self.at)
        self.ws.row_dimensions[self.at].height = 20
        self.at += 1

    def group(self, text):
        """그룹 머리 — 화면의 bucket-head(딜 소개 / 8월 넷째주 / IR 요청 투자사)."""
        from openpyxl.styles import Alignment, Font, PatternFill

        cell = self.ws.cell(row=self.at, column=1, value=text)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill("solid", fgColor=_R_GROUP_FILL)
        cell.alignment = Alignment(vertical="center")
        self._merge(self.at)
        self.at += 1

    def note(self, text, level=""):
        """한 줄 안내. **칸을 합치지 않는다.**

        합치면 엑셀이 줄 높이를 자동으로 늘려 주지 않아서, 폭보다 긴 글이
        접힌 채 첫 줄만 보인다. 안 나간 건을 알리는 줄이 그렇게 잘리면 이
        문서가 하려던 말이 통째로 사라진다. 안 합치면 오른쪽 빈 칸으로
        넘쳐 흘러 한 줄로 다 보인다.
        """
        from openpyxl.styles import Font

        color = {"bad": _R_BAD, "warn": _R_WARN}.get(level, _R_MUTED)
        cell = self.ws.cell(row=self.at, column=1, value=text)
        cell.font = Font(size=10, color=color, bold=level == "bad")
        self.at += 1

    def head(self, headers):
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        line = Side(style="thin", color=_R_LINE)
        lines = 1
        for i, text in enumerate(headers, start=1):
            cell = self.ws.cell(row=self.at, column=i, value=text)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill("solid", fgColor=_R_HEAD_FILL)
            # 칸 폭보다 긴 이름은 접는다. 잘리면 무엇을 세는 칸인지 안 보인다.
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = Border(top=line, bottom=line, left=line, right=line)
            width = self.ws.column_dimensions[self._letter(i)].width or 8
            lines = max(lines, -(-_report_width(text) // max(int(width) - 1, 1)))
        # 줄 높이를 **접힌 줄 수에 맞춰** 잡는다. 22 로 고정해 두었더니 요약의
        # `진행한 미팅` 이 두 줄로 접히면서 아랫줄이 잘렸다 — 무엇을 센 숫자인지
        # 모르는 표가 된다. 높이를 아예 안 잡는 방법도 있지만, 그러면 여는
        # 프로그램마다 다르게 그려진다.
        self.ws.row_dimensions[self.at].height = 8 + 13.5 * lines
        self.at += 1

    def row(self, values, level="", nums=(), wraps=()):
        """표 한 줄. `nums` 는 오른쪽 정렬할 칸(0부터), `wraps` 는 접을 칸."""
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        line = Side(style="thin", color=_R_LINE)
        fill = PatternFill("solid", fgColor=_R_BAD_FILL) if level == "bad" else None
        color = {"bad": _R_BAD, "warn": _R_WARN}.get(level)
        for i, value in enumerate(values, start=1):
            cell = self.ws.cell(row=self.at, column=i,
                                value="" if value is None else value)
            cell.font = Font(size=10, color=color) if color else Font(size=10)
            if fill is not None:
                cell.fill = fill
            cell.alignment = Alignment(
                horizontal="center" if (i - 1) in nums else "left",
                vertical="top" if (i - 1) in wraps else "center",
                wrap_text=(i - 1) in wraps)
            cell.border = Border(bottom=line, left=line, right=line)
        self.at += 1

    def stats(self, pairs):
        """`라벨 · 값` 목록 — 화면의 import-stats(미팅 결과 · IR 자료 요청)."""
        from openpyxl.styles import Alignment, Font

        for label, value in pairs:
            self.ws.cell(row=self.at, column=1, value=label).font = Font(size=10)
            cell = self.ws.cell(row=self.at, column=2, value=value)
            cell.font = Font(size=10, bold=True)
            cell.alignment = Alignment(horizontal="left")
            self.at += 1


def _kpi(sheet, data):
    """화면 맨 위 KPI 줄을 그대로. 가로 한 줄이라 칸 폭에 매이지 않는다."""
    sends = data["sends"]
    labels = ["보낸 건수", "발송 회차", "안 나감", "잡은 미팅", "진행한 미팅",
              "결과 물어봄", "아직 안 물어봄", "IR 요청", "IR 전달"]
    values = [sends["sent"], sends["rounds"], sends["left"], data["total"],
              data["done"], data["followup_done"], data["followup_open"],
              data["ir_requested"], data["ir_delivered"]]
    sheet.head(labels)
    sheet.row(values, nums=set(range(len(values))))


def _sends_sheet(sheet, data, team_wide):
    """발송 — 화면 맨 위 패널. 카톡으로 손으로 쓰던 보고가 이 표다."""
    sends = data["sends"]
    sheet.band(f"{data['month']}월 발송  ·  딜 소개 · 딜 소싱")
    if sends["left"]:
        # 화면이 먼저 말하는 것과 같은 말. 이 줄이 없으면 대상 수를 완료로
        # 옮겨 적는 일이 그대로 다시 일어난다.
        sheet.note(f"대상이었는데 {sends['left']}건이 안 나갔습니다"
                   f"(회차 {sends['short']}개) — 아래 '완료'는 실제로 도착한 건만"
                   f" 셉니다. 대상 수를 완료로 적지 마세요.", level="bad")
    if not sends["rounds"]:
        sheet.note("이 달에는 나간 회차가 없습니다.")
        return

    # 딜 소개 칸은 **두 그룹 모두에 세운다.** 화면은 소싱에서 이 칸을 빼는데
    # (늘 0 인 칸이 눈을 잡으므로), 엑셀은 칸 폭이 시트 단위라 그룹마다 칸이
    # 어긋나면 두 표가 서로 다른 자리에 서 버린다.
    headers = ["날짜", "회차명", "딜 소개", "대상", "완료", "안 나감",
               "안 나간 사유", "상태"] + (["팀원"] if team_wide else [])
    nums = {2, 3, 4, 5}
    for group in sends["groups"]:
        sheet.blank()
        head = (f"{group['label']}   {group['sent']}건 완료"
                f"   ·   회차 {group['rounds']}개 · 대상 {group['contacts']}명")
        if group["companies"]:
            head += f" · 딜 소개 {group['companies']}개사"
        sheet.group(head)
        if not group["rows"]:
            sheet.note("이 달에는 없습니다.")
            continue
        sheet.head(headers)
        for r in group["rows"]:
            sheet.row(
                [r["day"] or "-", r["title"], r["companies"], r["target"],
                 r["sent"], r["left"], r["left_label"], r["status_label"]]
                + ([r["owner"]] if team_wide else []),
                level=r["level"], nums=nums, wraps={6})
        # 합계는 **건을 더한 값**이다. 위 그룹 머리의 '대상 N명' 과 다를 수 있는데,
        # 그쪽은 겹치는 사람을 한 명으로 센 값이라 그렇다(회차 둘의 대상이 겹친다).
        #
        # 딜 소개(개사) 칸은 **비워 둔다.** 회차마다 같은 딜을 다시 돌리므로
        # 더하면 같은 기업을 여러 번 센다 — 8개사 회차를 다시 돌린 달이 16개사가
        # 된다. 그 달에 소개한 기업 수는 위 그룹 머리가 겹치지 않게 세어 말한다.
        sheet.row(["합계", "", "", group["target"], group["sent"], group["left"],
                   "", ""] + ([""] if team_wide else []),
                  level="warn" if group["left"] else "", nums=nums)


def _meetings_sheet(sheet, data, team_wide):
    """미팅 — 주차별. 화면의 '{달}월 미팅 총 N개사' 패널."""
    sheet.title(f"{data['year']}년 {data['month']}월 미팅  총 {data['total']}개사",
                f"결과 문의는 미팅 뒤 {data['followup_days']}일쯤 겁니다 — "
                f"그걸 놓치면 계약이 흐지부지됩니다.")
    headers = (["날짜", "담당자", "투자사", "기업", "구분"]
               + (["팀원"] if team_wide else [])
               + ["결과", "결과 문의", "후기 · 들은 내용"])
    last = len(headers) - 1
    if not data["weeks"]:
        sheet.blank()
        sheet.note("이 달에는 기록된 미팅이 없습니다.")
    for week in data["weeks"]:
        sheet.blank()
        sheet.group(f"{week['label']}   {week['done']}개사 완료")
        sheet.head(headers)
        for m in week["items"]:
            # 화면의 '결과 문의' 칸과 **같은 말**이어야 한다. 여기서 따로 지어내면
            # 같은 건이 화면과 파일에서 다르게 불린다.
            if m["status"] != "done":
                ask = "미완료"
            elif m["followup_done"]:
                ask = "완료"
            elif not m["needs_followup"]:
                ask = "문의 불필요"
            elif m["followup_late"]:
                ask = f"{m['followup_due']} 지남"
            else:
                ask = f"{m['followup_due']} 예정"
            told = [t for t in (m["note"], (f"문의: {m['followup_note']}"
                                            if m["followup_note"] else "")) if t]
            sheet.row(
                [m["date"], m["name"], m["firm"], m["company"], m["kind"]]
                + ([m["owner"]] if team_wide else [])
                + [m["outcome"] or "-", ask, "\n".join(told)],
                level="bad" if m["followup_late"] else "", wraps={last})

    sheet.blank()
    sheet.band("미팅 결과")
    sheet.stats(data["outcomes"] or [("완료된 미팅이 없습니다.", "")])
    if data["canceled"]:
        sheet.note(f"취소 {data['canceled']}건")
    sheet.blank()
    sheet.band(f"IR 자료 요청  ·  {data['month']}월")
    sheet.stats([("요청받음", data["ir_requested"]),
                 ("전달함", data["ir_delivered"]),
                 ("아직 안 보냄", data["ir_open"])])


def _buckets_sheet(sheet, data, team_wide):
    """이 달의 반응 다섯 갈래 — 이름과 날짜를 나란히 둔다."""
    sheet.title(f"{data['year']}년 {data['month']}월 · 이 달의 반응",
                "숫자만 보면 '그게 누구였지' 가 이어집니다 — 이름과 날짜를 함께 둡니다.")
    headers = (["날짜", "담당자", "직함", "투자사", "기업", "상태"]
               + (["담당 팀원"] if team_wide else []))
    for bucket in data["buckets"]:
        sheet.blank()
        sheet.group(f"{bucket['label']}   {len(bucket['rows'])}건")
        if not bucket["rows"]:
            sheet.note("없습니다.")
            continue
        sheet.head(headers)
        for r in bucket["rows"]:
            sheet.row([r["date"], r["name"], r["title"], r["firm"],
                       r["company"], r["note"]]
                      + ([r["owner"]] if team_wide else []))


def report_workbook(data: dict, *, team_wide: bool, who: str,
                    today: date) -> bytes:
    """업무 보고 한 달치 → 리포트 엑셀 바이트.

    `data` 는 화면이 쓰는 것과 **같은 dict**(`report.monthly`)다. 여기서 다시
    세지 않는다 — 두 곳에서 세면 화면과 파일의 숫자가 갈리고, 그러면 어느 쪽을
    믿을지 알 수 없다.
    """
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - 배포 이미지에는 항상 있다
        raise HTTPException(status_code=500,
                            detail="엑셀 쓰기 모듈(openpyxl)이 설치되지 않았습니다.")

    year, month = data["year"], data["month"]
    tag = f"{year}-{month:02d}"
    wb = openpyxl.Workbook()
    # 되올리기 방지 표식 — 다른 내려받기와 같다. 보고 파일을 '투자사 관리 현황'
    # 업로드 칸에 잘못 넣으면 활동 이력이 뻥튀기된다.
    wb.properties.keywords = sp.EXPORT_MARK

    ws = wb.active
    ws.title = f"{tag} 발송"
    sends = _ReportSheet(ws, [12, 34, 10, 9, 9, 10, 30, 15, 10],
                         9 if team_wide else 8)
    sends.title(f"{year}년 {month}월 업무 보고",
                f"{who} · {today.isoformat()} 뽑음 · "
                f"'완료'는 실제로 도착한 건만 셉니다(중단된 회차는 완료가 아닙니다).")
    sends.blank()
    sends.band("요약")
    _kpi(sends, data)
    sends.blank()
    _sends_sheet(sends, data, team_wide)

    _meetings_sheet(
        _ReportSheet(wb.create_sheet(f"{tag} 미팅"),
                     [12, 14, 24, 20, 10, 10, 12, 18, 46],
                     9 if team_wide else 8),
        data, team_wide)
    _buckets_sheet(
        _ReportSheet(wb.create_sheet(f"{tag} 반응"),
                     [12, 14, 12, 26, 26, 16, 10],
                     7 if team_wide else 6),
        data, team_wide)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


@router.get("/api/export/report.xlsx")
def export_report(month: str = "", scope: str = "", member: int = 0,
                  db: Session = Depends(get_db),
                  user: User = Depends(get_current_user)):
    """업무 보고 화면 그대로 → 엑셀 리포트.

    주소의 `month`·`scope`·`member` 는 화면과 같은 것을 읽는다
    (`report.parse_month`·`report.scope_for`). 같은 주소에서 화면과 파일이 다른
    달·다른 범위를 내면 안 된다.
    """
    today = date.today()
    year, mon = report_svc.parse_month(month, today)
    who, team_wide, viewing = report_svc.scope_for(db, user, scope, member)
    data = report_svc.monthly(db, year, mon, who, today)
    content = report_workbook(
        data, team_wide=team_wide,
        who=("팀 전체" if team_wide else f"{viewing.name} 담당"), today=today)
    return Response(content=content, media_type=sp.XLSX_MEDIA_TYPE,
                    headers=sp.content_disposition(
                        f"업무보고_{year}-{mon:02d}.xlsx"))
