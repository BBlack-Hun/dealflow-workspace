"""엑셀·CSV 를 읽고 쓴다.

업로드는 팀원마다 관리하는 '투자사 관리 현황' 시트를 그대로 올리는 용도다.
구글 시트에서 내려받으면 .xlsx 이고, 가끔 .csv 로 내보내는 사람도 있어 둘 다 받는다.
읽은 결과는 기존 임포트 파서(`sheet_import.parse_sheet_a`)가 쓰는 모양,
즉 **문자열 2차원 배열**로 맞춘다 — 파서를 새로 만들지 않기 위해서다.

내려받기는 화면의 표를 그대로 엑셀로 준다. 숫자는 숫자로 쓴다(엑셀에서 합계·정렬이
되어야 한다). 다만 휴대폰번호처럼 0 으로 시작하는 값은 문자열로 둔다.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Sequence

# 업로드 상한. 팀 시트는 수백 행이라 이보다 훨씬 작다. 큰 파일로 메모리를 먹는 것을 막는다.
MAX_UPLOAD_BYTES = 10 * 1024 * 1024
EXCEL_SUFFIXES = (".xlsx", ".xlsm")
CSV_SUFFIXES = (".csv", ".tsv", ".txt")


class SpreadsheetError(ValueError):
    """사용자에게 그대로 보여줄 수 있는 실패 사유."""


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def sheet_names(filename: str, data: bytes) -> List[str]:
    """엑셀이면 시트 이름 목록, CSV 면 빈 목록."""
    if not _is_excel(filename):
        return []
    wb = _load_workbook(data)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_rows(filename: str, data: bytes, sheet: Optional[str] = None) -> List[List[str]]:
    """업로드 파일 → 문자열 2차원 배열.

    엑셀은 시트를 고를 수 있고(없으면 첫 시트), CSV 는 시트 개념이 없다.
    """
    if len(data) > MAX_UPLOAD_BYTES:
        raise SpreadsheetError(
            f"파일이 너무 큽니다 ({len(data) // 1024 // 1024}MB). "
            f"{MAX_UPLOAD_BYTES // 1024 // 1024}MB 이하로 올려주세요."
        )
    if not data:
        raise SpreadsheetError("빈 파일입니다.")

    if _is_excel(filename):
        return _read_excel(data, sheet)
    if filename.lower().endswith(CSV_SUFFIXES):
        return _read_csv(data)
    raise SpreadsheetError(
        f"읽을 수 없는 형식입니다: {filename}. .xlsx 또는 .csv 로 올려주세요."
    )


def _is_excel(filename: str) -> bool:
    return filename.lower().endswith(EXCEL_SUFFIXES)


def _load_workbook(data: bytes):
    try:
        import openpyxl
    except ImportError:  # pragma: no cover - 배포 이미지에는 항상 있다
        raise SpreadsheetError("엑셀 읽기 모듈(openpyxl)이 설치되지 않았습니다.")
    try:
        # data_only=True → 수식이 아니라 계산된 값을 읽는다(시트에 수식이 많다).
        return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise SpreadsheetError(f"엑셀 파일을 열지 못했습니다: {exc}")


def _read_excel(data: bytes, sheet: Optional[str]) -> List[List[str]]:
    wb = _load_workbook(data)
    try:
        if sheet:
            if sheet not in wb.sheetnames:
                raise SpreadsheetError(
                    f"'{sheet}' 시트가 없습니다. 있는 시트: {', '.join(wb.sheetnames)}"
                )
            ws = wb[sheet]
        else:
            ws = wb[wb.sheetnames[0]]
        return [[_cell_to_text(c) for c in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _read_csv(data: bytes) -> List[List[str]]:
    # 구글 시트는 UTF-8(BOM), 국내 엑셀은 CP949 로 내보내는 일이 많다.
    for encoding in ("utf-8-sig", "cp949", "utf-8"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise SpreadsheetError("문자 인코딩을 알 수 없습니다. UTF-8 로 저장해 다시 올려주세요.")
    return [[(c or "").strip() for c in row] for row in csv.reader(io.StringIO(text))]


# --- 내려받기 ---------------------------------------------------------------

# 이 컬럼들은 숫자로 보여도 문자열로 써야 한다(앞의 0 이 사라지면 못 쓰는 값이 된다).
TEXT_COLUMNS = ("휴대폰", "연락처", "전화", "번호")


def _looks_numeric(value: str) -> bool:
    try:
        float(value.replace(",", ""))
    except (TypeError, ValueError):
        return False
    return not value.startswith("0") or value in ("0", "0.0")


# 내보낸 파일에 남기는 표식. 이 파일을 그대로 업로드하면 활동 이력이 뻥튀기된다
# (내보내기의 'IR 요청(누적)' 류 컬럼을 임포트 파서가 월별 활동으로 읽는다).
# 업로드에서 이 표식을 보고 막는다.
EXPORT_MARK = "dealflow-export"


def is_export_file(filename: str, data: bytes) -> bool:
    """우리가 내보낸 파일인지."""
    if not _is_excel(filename):
        return False
    try:
        wb = _load_workbook(data)
    except SpreadsheetError:
        return False
    try:
        return (wb.properties.keywords or "") == EXPORT_MARK
    finally:
        wb.close()


def write_xlsx(sheet_title: str, headers: Sequence[str],
               rows: Sequence[Sequence[Any]]) -> bytes:
    """표 하나를 담은 .xlsx 바이트. 머리행 고정 + 컬럼 폭 자동."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:  # pragma: no cover
        raise SpreadsheetError("엑셀 쓰기 모듈(openpyxl)이 설치되지 않았습니다.")

    wb = openpyxl.Workbook()
    wb.properties.keywords = EXPORT_MARK   # 되올리기 방지 표식
    ws = wb.active
    # 엑셀 시트 이름 제한: 31자, : \ / ? * [ ] 금지
    ws.title = _safe_sheet_title(sheet_title)

    ws.append(list(headers))
    head_fill = PatternFill("solid", fgColor="EEF2F7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")

    text_cols = {i for i, h in enumerate(headers)
                 if any(t in str(h) for t in TEXT_COLUMNS)}

    for row in rows:
        out = []
        for i, value in enumerate(row):
            if value is None:
                out.append("")
            elif isinstance(value, (int, float)):
                out.append(value)
            elif i not in text_cols and isinstance(value, str) and _looks_numeric(value):
                out.append(float(value.replace(",", "")))
            else:
                out.append(str(value))
        ws.append(out)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for i, header in enumerate(headers, start=1):
        longest = max(
            [_display_width(str(header))]
            + [_display_width(str(r[i - 1])) for r in rows[:200] if i - 1 < len(r)]
            or [8]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 8), 55)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _safe_sheet_title(title: str) -> str:
    for ch in ":\\/?*[]":
        title = title.replace(ch, " ")
    return (title.strip() or "Sheet")[:31]


def _display_width(text: str) -> int:
    """한글은 폭이 두 배라 그만큼 세어야 컬럼이 안 잘린다."""
    return sum(2 if ord(ch) > 0x2000 else 1 for ch in text)


def content_disposition(filename: str) -> Dict[str, str]:
    """한글 파일명이 깨지지 않게 RFC 5987 형식으로 넣는다."""
    from urllib.parse import quote

    return {"Content-Disposition":
            f"attachment; filename=\"download.xlsx\"; filename*=UTF-8''{quote(filename)}"}


XLSX_MEDIA_TYPE = ("application/vnd.openxmlformats-officedocument"
                   ".spreadsheetml.sheet")
