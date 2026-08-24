"""스타트업DB(기업정보) 시트를 가져온다.

시트의 `스타트업DB(기업정보)` 탭에는 대표자·연락처·이메일·연도별 매출·설립년도·
보증기관이 들어 있는데, 앱에는 그 칸만 있고 값이 없었다. 화면을 열면 텅 비어
보인다.

## 금액은 **적은 그대로** 담는다

원본 한 칸에 `8.2억` · `1,224백만원` · `150억 ~ 200억` 이 섞여 있다. 숫자로
바꾸려면 단위를 판별해야 하는데, 잘못 읽으면 100배가 틀어진 채 딜소개 문구에
실려 나간다. 사람이 적은 것이 곧 사실이므로 글자 그대로 둔다.

## 덮어쓰지 않는다

이미 값이 있는 칸은 건드리지 않는다(`--overwrite` 로만 덮는다). 앱에서 고친
내용이 임포트 한 번에 사라지면 안 된다.

    python scripts/import_startup_db.py 파일.xlsx            # 미리보기만
    python scripts/import_startup_db.py 파일.xlsx --apply
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.db import SessionLocal  # noqa: E402
from app.models import IrCompany  # noqa: E402

# 계약여부는 시트에 적힌 말 → 저장하는 값
CONTRACT_FROM_SHEET = {
    "미계약": "none", "무료계약완료": "free", "유료계약완료": "paid",
    "계약검토중": "review", "딜소개 불가": "blocked", "딜소개불가": "blocked",
}

# IR 기업현황(진행관리) 탭 — 시트 컬럼 → 모델 칸
STATUS_COLUMNS = [
    ("사업분야 대분류", "sector_major"),
    ("소분류", "sector_minor"),
    ("기업구분", "series"),
    ("한줄 소개", "one_liner"),
    ("담당자", "assignee_name"),
    ("계약여부", "contract_status"),
    ("날짜 기입", "contract_month"),
    ("핵심/TOP Deal", "top_deal_kind"),
]

# 시트 컬럼 → 모델 칸. 이름이 조금씩 달라도 찾도록 '포함' 으로 맞춘다.
COLUMNS = [
    ("기업명", None),
    ("대표자", "contact_name"),
    ("연락처", "contact_phone"),
    ("이메일", "contact_email"),
    # 시트의 `사업분야` 는 카테고리가 아니라 **사업 설명**이다.
    # 카테고리(대분류/소분류)는 IR 기업현황 탭이 따로 들고 있다.
    ("사업분야", "business_desc"),
    ("22년 매출", "revenue_2022"),
    ("23년 매출", "revenue_2023"),
    ("24년 매출", "revenue_2024"),
    ("25년 매출", "revenue_2025"),
    ("특이사항", "competitiveness"),
    ("설립년도", "founded_year"),
    ("기보", "guarantee"),
]


def norm(name) -> str:
    return re.sub(r"\(주\)|주식회사|㈜|\s", "", str(name or "")).strip()


def text(value) -> str:
    """셀을 글자로. 숫자는 그대로 찍되 불필요한 `.0` 은 뗀다."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def find_sheet(book, prefix):
    for name in book.sheetnames:
        if name.strip().startswith(prefix):
            return book[name]
    raise SystemExit(f"'{prefix}…' 탭을 찾을 수 없습니다: {book.sheetnames}")


def load(ws, columns, by_name, args, create=False):
    """한 탭을 읽어 채운다. (채운 칸 수, 찾은 기업 수, 못 찾은 이름, 새로 만든 수)"""
    head = {text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    where = {}
    for label, field in columns:
        col = next((c for name, c in head.items() if label in name), None)
        if col is None:
            print(f"  ! '{label}' 칸이 시트에 없습니다 — 건너뜁니다")
            continue
        where[label] = (col, field)

    name_col = next((c for name, c in head.items() if "기업명" in name), None)
    if name_col is None:
        raise SystemExit("기업명 칸을 찾을 수 없습니다")

    filled = matched = made = 0
    unmatched = []
    for r in range(2, ws.max_row + 1):
        raw = text(ws.cell(r, name_col).value)
        key = norm(raw)
        if not key:
            continue
        company = by_name.get(key)
        if company is None:
            if not create:
                unmatched.append(raw)
                continue
            # 시트에 새로 생긴 기업. 이름만이라도 만들어 둬야 다음 탭이 붙는다.
            company = IrCompany(name=raw, summary_status="draft")
            by_name[key] = company
            made += 1
        matched += 1
        for label, (col, field) in where.items():
            if field is None:
                continue
            value = text(ws.cell(r, col).value)
            if not value:
                continue
            if field == "contract_status":
                value = CONTRACT_FROM_SHEET.get(value, value)
            if field == "top_deal_kind":
                # 골라 넣으면 '추천 딜' 도 함께 켜진다
                company.is_top_deal = 1
            if getattr(company, field) and not args.overwrite:
                continue        # 앱에서 고친 내용을 임포트가 지우면 안 된다
            setattr(company, field, value)
            filled += 1
    return filled, matched, unmatched, made


def main() -> None:
    ap = argparse.ArgumentParser(description="IR 기업현황 · 스타트업DB 시트 가져오기")
    ap.add_argument("path")
    ap.add_argument("--apply", action="store_true", help="실제로 저장")
    ap.add_argument("--overwrite", action="store_true",
                    help="이미 값이 있는 칸도 덮어쓴다")
    args = ap.parse_args()

    book = openpyxl.load_workbook(args.path)
    db = SessionLocal()
    by_name = {norm(c.name): c
               for c in db.execute(select(IrCompany)).scalars().all()}

    # ① IR 기업현황이 먼저다 — 여기 있는 기업이 곧 대상 목록이라,
    #    없는 기업은 만들어 둬야 스타트업DB 의 기초자료가 붙을 자리가 생긴다.
    print("── IR 기업현황(진행관리) ──")
    f1, m1, u1, made = load(find_sheet(book, "IR 기업현황"),
                            STATUS_COLUMNS, by_name, args, create=True)
    for company in by_name.values():
        if company not in db:
            db.add(company)
    db.flush()
    print(f"  기업 {m1}개 · 새로 만듦 {made}개 · 채운 칸 {f1}개")

    print("── 스타트업DB(기업정보) ──")
    # 스타트업DB 에만 있는 기업도 만든다 — 기초자료가 있는데 자리가 없어서
    # 버리면, 나중에 딜소개 대상으로 올릴 때 다시 손으로 넣어야 한다.
    f2, m2, _u2, made2 = load(find_sheet(book, "스타트업DB"),
                              COLUMNS, by_name, args, create=True)
    for company in by_name.values():
        if company not in db:
            db.add(company)
    print(f"  기업 {m2}개 · 새로 만듦 {made2}개 · 채운 칸 {f2}개")

    if args.apply:
        db.commit()
        print("→ 저장했습니다.")
    else:
        db.rollback()
        print("→ 미리보기입니다. 실제로 넣으려면 --apply 를 붙이세요.")
    db.close()


if __name__ == "__main__":
    main()
