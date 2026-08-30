"""투자컨설턴트 현황 시트 가져오기.

원본이 `스타트업` · `경영본부 전달 기업` 으로 나뉘어 있고 관리하는 사람이
다르다. 시트별로 통째로 갈아 끼운다 — 맞춰 넣으면 시트에서 지운 줄이 앱에 남는다.

컬럼 순서가 시트마다 다르다(`경영본부 전달 기업` 은 기업명이 뒤에 있다).
자리가 아니라 **이름으로** 찾는다.

월 컬럼(`8월 마지막주 리마인드 …`)도 시트마다 달라서 시트별로 만든다 —
섞으면 없는 달의 빈 칸이 생긴다.

    python scripts/import_consulting.py 파일.xlsx            # 미리보기
    python scripts/import_consulting.py 파일.xlsx --apply
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from sqlalchemy import delete  # noqa: E402

from app.db import SessionLocal  # noqa: E402
from app.models import ConsultingColumn, ConsultingCompany  # noqa: E402
from app.routers.consulting import split_contract_line  # noqa: E402

# 시트 컬럼(포함으로 찾는다) → 모델 칸
FIELDS = [
    ("지역", "region"),
    ("미팅일", "meeting_at"),
    ("기업명", "company_name"),
    ("기업 관리", "management"),
    ("대표자", "ceo_name"),
    ("연락처", "phone"),
    ("이메일", "email"),
]
# 월 컬럼은 이 낱말이 들어간 칸으로 알아본다.
MONTH_MARK = "리마인드"

# 원본 파일의 시트 이름 → 앱의 탭 이름.
#
# 첫 탭 이름을 `중요 스타트업` 에서 `스타트업` 으로 바꿨는데(0039 마이그레이션),
# 사람이 들고 있는 xlsx 는 여전히 옛 이름이다. 그대로 넣으면 **옛 이름의 유령
# 탭이 다시 생기고** 그때부터 같은 명단이 두 탭으로 갈린다. 파일을 고치라고
# 하는 것보다 여기서 받아 주는 편이 안전하다.
SHEET_ALIAS = {"중요 스타트업": "스타트업"}


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    out = str(value).replace("\r", "").strip()
    # 구글 시트가 좁은 칸을 `#####` 로 내보낸다 — 값이 아니라 화면 표시다.
    if out and set(out) == {"#"}:
        return ""
    # `=ROW()-3` 같은 수식은 값이 아니다.
    return "" if out.startswith("=") else out


# `월간 계약 업무현황표` 는 다른 두 시트와 모양이 다르다. 머리글이 있는 표가
# 아니라 **월 묶음 + 슬래시 한 줄**이다:
#
#     6월  (무료계약 2개사 / 유료계약 3개사)
#     기업명 / 계약금액 / 성공보수율 / 계약일
#     ○○○/ 무료/ 3.5%/ 미정
#
# 이 모양 때문에 처음엔 건너뛰었는데, 그러면 화면에서 이 표를 아예 볼 수 없다.
#
# 처음에는 슬래시 줄을 `company_name` 에 통째로 넣었다. 한 칸에 뭉쳐 있으면
# 계약금으로 거를 수도, 보수율만 고칠 수도 없어서 **시트가 적어 둔 머리글
# 순서대로 칸에 나눠 담는다**(`기업명 / 계약금액 / 성공보수율 / 계약일`).
# 나누는 규칙은 앱과 같은 것을 쓴다 — 여기 따로 적으면 다시 올릴 때마다 화면과
# 다른 모양이 들어간다. 나누기 전 줄은 `source_line` 에 그대로 남는다.
CONTRACT_SHEET = "월간 계약 업무현황표"
_MONTH_LINE = re.compile(r"^\s*(\d{1,2})\s*월")
_HEADER_LINE = re.compile(r"기업명\s*/")


def parse_contract_sheet(ws) -> list:
    """월 묶음 자유 서식 → 줄 목록."""
    out, month = [], ""
    for r in range(1, ws.max_row + 1):
        label = text(ws.cell(r, 1).value)
        body = text(ws.cell(r, 2).value)

        m = _MONTH_LINE.match(label) or _MONTH_LINE.match(body)
        if m and "/" not in (label + body).replace(m.group(0), "", 1)[:3]:
            month = f"{int(m.group(1))}월"
            continue
        if not body or _HEADER_LINE.search(body):
            continue          # 머리글 줄은 값이 아니다
        if "/" not in body:
            continue          # 계약 줄이 아니다

        # 무료·유료는 줄 안에 적혀 있다. 왼쪽 라벨은 병합 때문에 줄과
        # 어긋나 있어(3행이 '무료 계약', 4행이 '유료 계약') 믿을 수 없다.
        kind = "유료" if "유료" in body else ("무료" if "무료" in body else "")
        out.append({"month": month, "kind": kind, "line": body})
    return out


def header_row(ws) -> int:
    """머리글 행. 시트마다 4~5행이다(위에 제목·요약이 붙어 있다)."""
    for r in range(1, 8):
        labels = [text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if sum(1 for x in labels if x) >= 4 and any("기업" in x for x in labels):
            return r
    return 1


def main() -> None:
    ap = argparse.ArgumentParser(description="투자컨설턴트 현황 가져오기")
    ap.add_argument("path")
    ap.add_argument("--apply", action="store_true", help="실제로 저장")
    ap.add_argument("--owner", type=int, default=0,
                    help="이 표의 주인(users.id). 투자컨설턴트 현황은 사람별이다")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.path)
    db = SessionLocal()
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet = sheet_name.strip()
        sheet = SHEET_ALIAS.get(sheet, sheet)
        # 계약 현황표는 머리글 있는 표가 아니라 월 묶음 자유 서식이다.
        if CONTRACT_SHEET in sheet:
            lines = parse_contract_sheet(ws)
            db.execute(delete(ConsultingCompany)
                       .where(ConsultingCompany.sheet == sheet))
            db.execute(delete(ConsultingColumn)
                       .where(ConsultingColumn.sheet == sheet))
            for pos, item in enumerate(lines, start=1):
                parts = split_contract_line(item["line"])
                db.add(ConsultingCompany(
                    sheet=sheet, position=pos, user_id=args.owner or None,
                    region=item["month"],            # 어느 달의 계약인가
                    management=item["kind"],         # 계약여부 — 무료 / 유료
                    # 나누기 전 한 줄. 나눈 결과가 틀렸을 때 여기서 다시 나눈다.
                    source_line=item["line"],
                    company_name=parts.get("company_name") or item["line"],
                    contract_fee=parts.get("contract_fee"),
                    success_fee=parts.get("success_fee"),
                    meeting_at=parts.get("meeting_at")))   # 계약일
            total += len(lines)
            print(f"  {sheet:22} {len(lines):3}건 (월 묶음)")
            continue

        hr = header_row(ws)
        head = [(c, text(ws.cell(hr, c).value)) for c in range(1, ws.max_column + 1)]
        if not any("기업" in h for _c, h in head):
            print(f"  건너뜀 (표가 아님): {sheet}")
            continue

        where = {}
        for label, field in FIELDS:
            col = next((c for c, h in head if label in h), None)
            if col is not None:
                where[field] = col
        month_cols = [(c, h) for c, h in head if MONTH_MARK in h]

        # 시트를 통째로 갈아 끼운다 — 맞춰 넣으면 지운 줄이 남는다.
        db.execute(delete(ConsultingCompany).where(ConsultingCompany.sheet == sheet))
        db.execute(delete(ConsultingColumn).where(ConsultingColumn.sheet == sheet))

        cols = []
        for pos, (c, label) in enumerate(month_cols):
            col = ConsultingColumn(sheet=sheet, label=label, position=pos,
                                   user_id=args.owner or None)
            db.add(col)
            cols.append((c, col))
        db.flush()

        n = 0
        for r in range(hr + 1, ws.max_row + 1):
            name = text(ws.cell(r, where["company_name"]).value) if "company_name" in where else ""
            if not name:
                continue
            row = ConsultingCompany(sheet=sheet, position=n + 1, company_name=name,
                                    user_id=args.owner or None)
            for field, c in where.items():
                if field == "company_name":
                    continue
                value = text(ws.cell(r, c).value)
                if value:
                    setattr(row, field, value)
            notes = {}
            for c, col in cols:
                value = text(ws.cell(r, c).value)
                if value:
                    notes[str(col.id)] = value
            row.notes = json.dumps(notes, ensure_ascii=False)
            db.add(row)
            n += 1
        total += n
        print(f"  {sheet:24} {n:3}개사 · 월 컬럼 {len(cols)}개")

    print(f"\n합계 {total}개사")
    if args.apply:
        db.commit()
        print("→ 저장했습니다.")
    else:
        db.rollback()
        print("→ 미리보기입니다. 실제로 넣으려면 --apply 를 붙이세요.")
    db.close()


if __name__ == "__main__":
    main()
