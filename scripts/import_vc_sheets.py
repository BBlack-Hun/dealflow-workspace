"""투자사 명단 시트에서 빠진 칸 채우기.

명단을 처음 옮길 때 이름·회사·연락처는 들어왔는데 **선호 투자분야**처럼
사람이 나중에 적어 넣은 칸이 빠져 있었다(시트 80건 → 앱 3건). 그 칸이
비면 "이 투자사가 뭘 보고 싶어 하는지" 를 화면에서 알 수 없다.

## 이름으로 맞춘다
시트마다 같은 사람이 다른 표기로 있을 수 있어(`홍길동` vs `홍길동 이사님`)
직함을 떼고 맞춘다. 회사까지 같아야 같은 사람으로 본다 — 이름만 보면
동명이인이 섞인다.

## 덮어쓰지 않는다
이미 값이 있는 칸은 건드리지 않는다(`--overwrite` 로만 덮는다). 앱에서
고친 내용이 임포트 한 번에 사라지면 안 된다.

    python scripts/import_vc_sheets.py 파일.xlsx            # 미리보기
    python scripts/import_vc_sheets.py 파일.xlsx --apply
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.db import SessionLocal  # noqa: E402
from app.models import VcContact  # noqa: E402

# 투자사 명단 탭만 본다. 참고 자료 탭은 `import_ref_sheets.py` 가 맡는다.
LIST_SHEETS = re.compile(r"\d+\s*\(|\d+명")

# 시트 컬럼 → 모델 칸. 이름이 조금씩 달라도 찾도록 '포함' 으로 맞춘다.
COLUMNS = [
    ("담당자", "assignee_name"),
    ("메모", "memo"),
    ("선호 투자", "sectors"),          # `선호 투자분야` · `선호 투자 분야`
    ("라운드 사이즈", "round_size"),
    ("휴대폰", "phone"),
    ("회사", "firm"),
    ("부서", "department"),
    ("직함", "title"),
    ("전자 메일", "email"),
    # 아래는 시트에 있는데 그동안 통째로 버려지던 값들.
    ("근무처 전화", "office_phone"),
    ("근무처 팩스", "office_fax"),
    ("근무지 주소", "address"),
    ("명함 등록일", "card_registered_at"),
    ("관심도", "interest_level"),
]

# 이름 뒤에 붙는 존칭·직함. 맞출 때만 떼고, 저장된 이름은 건드리지 않는다.
HONORIFIC = re.compile(
    r"\s*(대표이사님|대표이사|대표님|대표|부서장님|부서장|본부장님|본부장|"
    r"팀장님|팀장|실장님|실장|이사님|이사|전무님|전무|상무님|상무|"
    r"부장님|부장|차장님|차장|과장님|과장|대리님|대리|주임님|주임|"
    r"사원님|사원|심사역님|심사역|파트너님|파트너|매니저님|매니저|"
    r"수석님|수석|책임님|책임|선임님|선임|님)\s*$")


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    out = str(value).replace("\r", "").strip()
    # 구글 시트가 **좁은 칸을 `#####` 로 내보낸다.** 값이 아니라 화면 표시라
    # 그대로 담으면 명함 등록일 자리에 `###########` 이 들어간다.
    if out and set(out) == {"#"}:
        return ""
    return out


def bare_name(value) -> str:
    """`홍길동 이사님` → `홍길동`. 표기가 갈려도 같은 사람으로 맞추기 위해."""
    name = text(value)
    for _ in range(3):                      # `홍길동 이사 님` 처럼 겹친 경우
        stripped = HONORIFIC.sub("", name)
        if stripped == name:
            break
        name = stripped
    return re.sub(r"\s+", "", name)


def firm_key(value) -> str:
    return re.sub(r"\(주\)|주식회사|㈜|\s", "", text(value))


def header_row(ws) -> int:
    """머리글 행. 시트마다 1행이거나 2행이다."""
    for r in range(1, 4):
        labels = [text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any("휴대폰" in x or "이름" in x for x in labels):
            return r
    return 1


def _guess_name_column(ws, hr: int, head: dict):
    """머리글이 비어 있는 칸에서 이름 칸을 찾는다.

    `NO` 다음에 오는, 짧은 한글이 대부분인 칸. 시트를 손으로 고쳐 달라고
    하는 것보다 이쪽이 낫다 — 원본은 사람이 계속 쓰는 문서다.
    """
    blank_cols = [c for c in range(1, ws.max_column + 1)
                  if not text(ws.cell(hr, c).value)]
    best, best_score = None, 0
    for c in blank_cols:
        values = [text(ws.cell(r, c).value)
                  for r in range(hr + 1, min(hr + 21, ws.max_row + 1))]
        values = [v for v in values if v]
        if not values:
            continue
        korean_short = sum(1 for v in values
                           if len(v) <= 8 and re.fullmatch(r"[가-힣\s]+", v))
        score = korean_short / len(values)
        if score > best_score and score >= 0.6:
            best, best_score = c, score
    return best


def main() -> None:
    ap = argparse.ArgumentParser(description="투자사 명단 시트 가져오기")
    ap.add_argument("path")
    ap.add_argument("--apply", action="store_true", help="실제로 저장")
    ap.add_argument("--overwrite", action="store_true",
                    help="이미 값이 있는 칸도 덮어쓴다")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.path)
    db = SessionLocal()
    people = db.execute(select(VcContact)).scalars().all()

    # 이름+회사 → 사람. 회사가 비어 있으면 이름만으로도 찾을 수 있게 둘 다 건다.
    by_pair, by_name = {}, {}
    for p in people:
        key = bare_name(p.name)
        if not key:
            continue
        by_pair[(key, firm_key(p.firm))] = p
        by_name.setdefault(key, []).append(p)

    filled = matched = 0
    unmatched = []
    for name in wb.sheetnames:
        if not LIST_SHEETS.search(name):
            continue
        ws = wb[name]
        hr = header_row(ws)
        head = {text(ws.cell(hr, c).value): c for c in range(1, ws.max_column + 1)}
        name_col = next((c for h, c in head.items() if "이름" in h), None)
        firm_col = next((c for h, c in head.items() if h.strip() == "회사"), None)
        if name_col is None:
            # `30(15명 연결)` 은 이름 칸의 머리글이 비어 있다. 머리글이 없는
            # 칸 중 값이 사람 이름처럼 생긴 칸을 찾는다 — 시트를 손으로
            # 고쳐 달라고 하는 것보다 이쪽이 낫다.
            name_col = _guess_name_column(ws, hr, head)
        if name_col is None:
            print(f"  건너뜀 (이름 칸 없음): {name}")
            continue

        where = {}
        for label, field in COLUMNS:
            col = next((c for h, c in head.items() if label in h), None)
            if col is not None:
                where[label] = (col, field)

        print(f"── {name} (머리 {hr}행) ──")
        for r in range(hr + 1, ws.max_row + 1):
            key = bare_name(ws.cell(r, name_col).value)
            if not key:
                continue
            firm = firm_key(ws.cell(r, firm_col).value) if firm_col else ""
            person = by_pair.get((key, firm))
            if person is None:
                candidates = by_name.get(key, [])
                # 동명이인이면 회사가 맞아야만 고른다 — 아니면 건너뛴다.
                person = candidates[0] if len(candidates) == 1 else None
            if person is None:
                unmatched.append(text(ws.cell(r, name_col).value))
                continue
            matched += 1
            for _label, (col, field) in where.items():
                value = text(ws.cell(r, col).value)
                if not value:
                    continue
                if getattr(person, field) and not args.overwrite:
                    continue     # 앱에서 고친 내용을 임포트가 지우면 안 된다
                setattr(person, field, value)
                filled += 1

    print(f"\n찾은 사람 {matched}명 · 채운 칸 {filled}개")
    if unmatched:
        print(f"이름이 안 맞는 사람 {len(unmatched)}명: "
              + ", ".join(unmatched[:6]) + (" …" if len(unmatched) > 6 else ""))

    if args.apply:
        db.commit()
        print("→ 저장했습니다.")
    else:
        db.rollback()
        print("→ 미리보기입니다. 실제로 넣으려면 --apply 를 붙이세요.")
    db.close()


if __name__ == "__main__":
    main()
