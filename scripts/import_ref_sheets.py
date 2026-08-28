"""참고 시트 가져오기 — 원본 스프레드시트의 '자료' 탭들.

투자사 명단 말고도 시트에는 스크립트·가이드·성격 정리 같은 탭이 여럿 있었다.
매번 구글 시트를 따로 열어 보던 자료라 화면 안으로 들여온다.

## 모양이 제각각이다

    table — `투자사 성격정리` 처럼 머리글 + 행이 또렷한 진짜 표
    text  — `딜소개 스크립트` 처럼 한 칸에 줄글이 들어 있는 문서

칸이 두 개 이상인 행이 여러 줄 이어지면 표로, 아니면 줄글로 본다.
애매하면 **줄글로 둔다** — 표로 잘못 읽으면 내용이 칸에 잘려 사라지지만,
줄글로 두면 적어도 원문이 남는다.

## 투자사 명단 탭은 가져오지 않는다
`150(71명연결)` 같은 탭은 이미 투자사 관리 현황 본문이 다룬다. 두 벌로 두면
어느 쪽이 최신인지 알 수 없다.

    python scripts/import_ref_sheets.py 파일.xlsx            # 미리보기
    python scripts/import_ref_sheets.py 파일.xlsx --apply
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.db import SessionLocal  # noqa: E402
from app.models import RefSheet  # noqa: E402

# 명단 탭 — 본문이 이미 다룬다. 이름에 이런 꼴이 들어가면 건너뛴다.
#
# 예외로 둘 탭은 **부르는 사람이 `--keep` 으로 정한다.** 예전에는 탭 이름
# 하나를 여기 박아 두었는데, 그런 예외는 늘 하나로 끝나지 않는다 — 다음 탭에서
# 또 박아야 하고, 박는 것을 잊으면 그 탭만 조용히 사라진다.
SKIP = re.compile(r"\d+\s*\(|\d+명|전체 딜소개현황")


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).replace("\r", "").strip()


def read_grid(ws) -> list:
    """빈 행·빈 열을 걷어낸 격자."""
    grid = []
    for r in range(1, ws.max_row + 1):
        row = [text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any(row):
            grid.append(row)
    if not grid:
        return []
    width = max(len(r) for r in grid)
    keep = [c for c in range(width)
            if any(c < len(r) and r[c] for r in grid)]
    return [[r[c] if c < len(r) else "" for c in keep] for r in grid]


def looks_like_table(grid: list) -> bool:
    """머리글이 있고, 칸이 둘 이상인 행이 대부분이면 표로 본다.

    **머리글은 폭이 아니라 채워진 칸 수로 본다.** 폭만 보면 `딜소개 스크립트`
    처럼 한 칸에 제목만 있고 나머지가 빈 문서도 4칸짜리 표로 읽힌다.

    애매하면 줄글이다 — 표로 잘못 읽으면 내용이 칸에 잘려 사라지지만,
    줄글로 두면 적어도 원문이 남는다.
    """
    if len(grid) < 3:
        return False
    if sum(1 for c in grid[0] if c) < 2:
        return False                       # 머리글이 한 칸뿐이면 표가 아니다
    multi = sum(1 for row in grid[1:] if sum(1 for c in row if c) >= 2)
    return multi >= max(3, len(grid[1:]) // 2)


def as_text(grid: list) -> str:
    """격자를 줄글로. 한 행의 칸들은 줄바꿈으로 잇는다."""
    out = []
    for row in grid:
        cells = [c for c in row if c]
        if cells:
            out.append("\n".join(cells))
    return "\n\n".join(out)


def main() -> None:
    ap = argparse.ArgumentParser(description="참고 시트 가져오기")
    ap.add_argument("path")
    ap.add_argument("--apply", action="store_true", help="실제로 저장")
    ap.add_argument("--page", default="contacts",
                    help="붙일 화면: contacts | consulting")
    ap.add_argument("--keep", default="",
                    help="명단처럼 보여도 가져올 탭 이름들(쉼표로 구분)")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.path)
    db = SessionLocal()
    existing = {s.title: s for s in db.execute(select(RefSheet)).scalars().all()}

    keep = [k.strip() for k in args.keep.split(",") if k.strip()]
    made = updated = 0
    for pos, name in enumerate(wb.sheetnames):
        title = name.strip()
        if SKIP.search(title) and not any(k in title for k in keep):
            print(f"  건너뜀 (명단 탭): {title}")
            continue

        grid = read_grid(wb[name])
        if not grid:
            print(f"  건너뜀 (빈 시트): {title}")
            continue

        if looks_like_table(grid):
            kind = "table"
            content = {"columns": grid[0], "rows": grid[1:]}
            shape = f"표 {len(grid[0])}칸 x {len(grid) - 1}줄"
        else:
            kind = "text"
            content = {"body": as_text(grid)}
            shape = f"줄글 {len(content['body'])}자"

        row = existing.get(title)
        if row is None:
            db.add(RefSheet(title=title, kind=kind, position=pos, page=args.page,
                            content_json=json.dumps(content, ensure_ascii=False)))
            made += 1
            print(f"  새로 만듦: {title:36} {shape}")
        else:
            row.kind = kind
            row.page = args.page
            row.position = pos
            row.content_json = json.dumps(content, ensure_ascii=False)
            updated += 1
            print(f"  갱신:      {title:36} {shape}")

    print(f"\n새로 {made}개 · 갱신 {updated}개")
    if args.apply:
        db.commit()
        print("→ 저장했습니다.")
    else:
        db.rollback()
        print("→ 미리보기입니다. 실제로 넣으려면 --apply 를 붙이세요.")
    db.close()


if __name__ == "__main__":
    main()
