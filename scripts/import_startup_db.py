"""스타트업DB(기업정보) 시트를 가져온다.

시트의 `스타트업DB(기업정보)` 탭에는 대표자·연락처·이메일·연도별 매출·설립년도·
보증기관이 들어 있는데, 앱에는 그 칸만 있고 값이 없었다. 화면을 열면 텅 비어
보인다.

## 금액은 **적은 그대로** 담는다

원본 한 칸에 `8.2억` · `1,224백만원` · `150억 ~ 200억` 이 섞여 있다. 숫자로
바꾸려면 단위를 판별해야 하는데, 잘못 읽으면 100배가 틀어진 채 딜소개 문구에
실려 나간다. 사람이 적은 것이 곧 사실이므로 글자 그대로 둔다.

## 덮어쓰지 않는다

이미 값이 있는 칸은 건드리지 않는다(`--overwrite` 로만 덮는다). 앱에서 고친
내용이 임포트 한 번에 사라지면 안 된다.

    python scripts/import_startup_db.py 파일.xlsx            # 미리보기만
    python scripts/import_startup_db.py 파일.xlsx --apply
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.db import SessionLocal  # noqa: E402
from app.models import IrCompany  # noqa: E402

SHEET = "스타트업DB(기업정보)"

# 시트 컬럼 → 모델 칸. 이름이 조금씩 달라도 찾도록 '포함' 으로 맞춘다.
COLUMNS = [
    ("기업명", None),
    ("대표자", "contact_name"),
    ("연락처", "contact_phone"),
    ("이메일", "contact_email"),
    ("22년 매출", "revenue_2022"),
    ("23년 매출", "revenue_2023"),
    ("24년 매출", "revenue_2024"),
    ("25년 매출", "revenue_2025"),
    ("특이사항", "competitiveness"),
    ("설립년도", "founded_year"),
    ("기보", "guarantee"),
]


def norm(name) -> str:
    return re.sub(r"\(주\)|주식회사|㈜|\s", "", str(name or "")).strip()


def text(value) -> str:
    """셀을 글자로. 숫자는 그대로 찍되 불필요한 `.0` 은 뗀다."""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def find_sheet(book):
    for name in book.sheetnames:
        if name.strip().startswith("스타트업DB"):
            return book[name]
    raise SystemExit(f"'{SHEET}' 탭을 찾을 수 없습니다: {book.sheetnames}")


def main() -> None:
    ap = argparse.ArgumentParser(description="스타트업DB 시트 가져오기")
    ap.add_argument("path")
    ap.add_argument("--apply", action="store_true", help="실제로 저장")
    ap.add_argument("--overwrite", action="store_true",
                    help="이미 값이 있는 칸도 덮어쓴다")
    args = ap.parse_args()

    ws = find_sheet(openpyxl.load_workbook(args.path))
    head = {text(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}

    where = {}
    for label, field in COLUMNS:
        match = next((c for name, c in head.items() if label in name), None)
        if match is None:
            print(f"  ! 시트에 '{label}' 칸이 없습니다 — 건너뜁니다")
            continue
        where[label] = (match, field)

    db = SessionLocal()
    by_name = {norm(c.name): c for c in db.execute(select(IrCompany)).scalars().all()}

    filled, matched, unmatched = 0, 0, []
    for r in range(2, ws.max_row + 1):
        raw = ws.cell(r, where["기업명"][0]).value
        key = norm(raw)
        if not key:
            continue
        company = by_name.get(key)
        if company is None:
            unmatched.append(text(raw))
            continue
        matched += 1
        for label, (col, field) in where.items():
            if field is None:
                continue
            value = text(ws.cell(r, col).value)
            if not value:
                continue
            if getattr(company, field) and not args.overwrite:
                continue        # 앱에서 고친 내용을 임포트가 지우면 안 된다
            setattr(company, field, value)
            filled += 1

    print(f"시트 기업 {matched + len(unmatched)}개 · 앱에서 찾음 {matched}개")
    print(f"채울 칸 {filled}개")
    if unmatched:
        print(f"이름이 안 맞는 기업 {len(unmatched)}개:")
        for n in unmatched[:8]:
            print(f"   - {n}")
        if len(unmatched) > 8:
            print(f"   … 외 {len(unmatched) - 8}개")

    if args.apply:
        db.commit()
        print("→ 저장했습니다.")
    else:
        db.rollback()
        print("→ 미리보기입니다. 실제로 넣으려면 --apply 를 붙이세요.")
    db.close()


if __name__ == "__main__":
    main()
