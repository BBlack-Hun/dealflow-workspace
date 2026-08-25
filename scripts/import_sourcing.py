"""딜 소싱 참여 심사역 시트 가져오기.

시트가 **무엇을 찾는지**로 나뉘어 있다(시리즈 A 이상 · 개인 참여 · M&A ·
후속투자). 같은 사람이 여러 갈래에 들어갈 수 있으므로 갈래를 지우고 다시
넣는다 — 갈래별로 통째로 갈아 끼우는 편이 맞춰 넣는 것보다 안전하다.

    python scripts/import_sourcing.py 파일.xlsx            # 미리보기
    python scripts/import_sourcing.py 파일.xlsx --apply
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from sqlalchemy import delete, select  # noqa: E402

from app.db import SessionLocal  # noqa: E402
from app.models import SourcingContact  # noqa: E402

# 시트 컬럼(포함으로 찾는다) → 모델 칸
COLUMNS = [
    ("참여 요청일", "requested_at"),
    ("이름", "name"),
    ("담당자", "assignee_name"),
    ("휴대폰", "phone"),
    ("직함", "title"),
    ("이메일", "email"),
    ("메모", "memo"),
    ("회사", "firm"),
    ("딜 공유 방법", "share_method"),
    ("투자분야", "sectors"),
    ("라운드 사이즈", "round_size"),
    ("TIPS", "tips"),
    ("카톡으로 받은", "kakao_reply"),
    ("통화내용", "call_note"),
]


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    out = str(value).replace("\r", "").strip()
    # 구글 시트가 좁은 칸을 `#####` 로 내보낸다 — 값이 아니라 화면 표시다.
    return "" if out and set(out) == {"#"} else out


# 시트의 담당자 칸에는 이름 뒤에 요약이 붙어 있다 — `홍길동 (총 4명)`.
# 그대로 두면 같은 사람이 둘로 갈려서, 담당으로 거를 때 한쪽이 사라진다.
SUMMARY_TAIL = re.compile(r"\s*\(\s*총\s*\d+\s*명?\s*\)\s*$")


def assignee(value: str) -> str:
    return SUMMARY_TAIL.sub("", value).strip()


def header_row(ws) -> int:
    """머리글 행. 시트마다 1~2행이다."""
    for r in range(1, 5):
        labels = [text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if sum(1 for x in labels if x) >= 3 and any("이름" in x for x in labels):
            return r
    return 1


def main() -> None:
    ap = argparse.ArgumentParser(description="딜 소싱 시트 가져오기")
    ap.add_argument("path")
    ap.add_argument("--apply", action="store_true", help="실제로 저장")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.path)
    db = SessionLocal()
    made = 0

    for pos, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        bucket = sheet_name.strip()
        hr = header_row(ws)
        head = {text(ws.cell(hr, c).value): c for c in range(1, ws.max_column + 1)}

        where = {}
        for label, field in COLUMNS:
            col = next((c for h, c in head.items() if label in h), None)
            if col is not None:
                where[field] = col
        if "name" not in where:
            print(f"  건너뜀 (이름 칸 없음): {bucket}")
            continue

        # 갈래를 통째로 갈아 끼운다 — 맞춰 넣으면 지워진 사람이 남는다.
        db.execute(delete(SourcingContact).where(SourcingContact.bucket == bucket))

        n = 0
        for r in range(hr + 1, ws.max_row + 1):
            name = text(ws.cell(r, where["name"]).value)
            if not name:
                continue
            row = SourcingContact(bucket=bucket, position=pos * 1000 + n, name=name)
            for field, col in where.items():
                if field == "name":
                    continue
                value = text(ws.cell(r, col).value)
                if field == "assignee_name":
                    value = assignee(value)
                if value:
                    setattr(row, field, value)
            db.add(row)
            n += 1
        made += n
        print(f"  {bucket:36} {n:3}명")

    print(f"\n합계 {made}명")
    if args.apply:
        db.commit()
        print("→ 저장했습니다.")
    else:
        db.rollback()
        print("→ 미리보기입니다. 실제로 넣으려면 --apply 를 붙이세요.")
    db.close()


if __name__ == "__main__":
    main()
