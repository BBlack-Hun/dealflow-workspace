"""딜 소싱 제안 문구 가져오기 — 갈래별 초대장.

원본 시트에는 표 오른쪽 빈 칸에 실제로 보내는 스크립트가 적혀 있다. 갈래마다
호칭·개수·범위·딜 수신 메일이 다르고, 그중 **메일 주소는 코드에 둘 수 없어**
(저장소가 공개다) 여기서 DB 로만 넣는다.

한 통에 인사말까지 들어 있는 원문에서 인사 줄은 떼어낸다 — 인사는 인사말
템플릿이 붙이므로, 그대로 두면 인사가 두 번 나간다.

    python scripts/import_sourcing_scripts.py 파일.xlsx           # 미리보기
    python scripts/import_sourcing_scripts.py 파일.xlsx --apply
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.db import SessionLocal  # noqa: E402
from app.models import MessageTemplate, SourcingContact  # noqa: E402
from app.services import sourcing_msg  # noqa: E402

# 이 말이 들어 있는 칸이 참여 안내문이다.
MARKER = "Deal Sourcing 네트워크 참여"

# 인사 줄 — 문구가 아니라 인사말 템플릿의 몫이다.
GREETING = re.compile(r"^\s*안녕하세요.*$")
# 원문에는 안내문 아래에 다른 스크립트(감사 톡 등)가 `===` 로 이어 붙어 있다.
DIVIDER = re.compile(r"^\s*=====+\s*$")


def find_script(ws) -> str:
    """이 시트에서 가장 긴 참여 안내문 칸.

    같은 말이 여러 칸에 흩어져 있고 짧은 것은 메모다 — 제일 긴 것이 본문이다.
    """
    best = ""
    for row in ws.iter_rows():
        for cell in row:
            v = str(cell.value or "").strip()
            if MARKER in v and len(v) > len(best):
                best = v
    return best


def clean(raw: str) -> str:
    """인사 줄과 뒤에 붙은 다른 스크립트를 떼어낸다."""
    lines = raw.splitlines()
    # 안내문이 시작되는 곳부터
    start = next((i for i, ln in enumerate(lines) if MARKER in ln), 0)
    out = []
    for line in lines[start:]:
        if DIVIDER.match(line):
            break                      # 여기서부터는 다른 스크립트다
        if GREETING.match(line):
            continue                   # 인사는 인사말이 맡는다
        out.append(line.rstrip())
    # 앞뒤 빈 줄 정리
    while out and not out[0].strip():
        out.pop(0)
    while out and not out[-1].strip():
        out.pop()
    return "\n".join(out)


def match_bucket(sheet_name: str, buckets: list) -> str:
    """시트 탭 이름 ↔ DB 갈래 이름. 공백·표기가 조금씩 다르다."""
    def key(s: str) -> str:
        return re.sub(r"\s+", "", s or "").lower()

    want = key(sheet_name)
    for b in buckets:
        if key(b) == want:
            return b
    # 완전히 같지 않으면 한쪽이 다른 쪽을 품는지 본다
    for b in buckets:
        if key(b) in want or want in key(b):
            return b
    return ""


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.path)
    db = SessionLocal()
    buckets = [b for (b,) in db.execute(
        select(SourcingContact.bucket).distinct()).all()]
    if not buckets:
        print("딜 소싱 명단이 비어 있습니다 — import_sourcing.py 를 먼저 돌리세요.")
        return 1

    found, missing = [], []
    for name in wb.sheetnames:
        raw = find_script(wb[name])
        if not raw:
            continue
        bucket = match_bucket(name, buckets)
        if not bucket:
            missing.append(name)
            continue
        found.append((bucket, clean(raw)))

    # 스크립트가 없는 갈래는 뼈대로 채운다 — 문구가 비면 발송 화면에서
    # 왜 안 되는지 알 수 없다.
    have = {b for b, _ in found}
    for bucket in buckets:
        if bucket not in have:
            found.append((bucket, sourcing_msg.default_body(bucket)))

    for bucket, body in found:
        mark = "시트" if bucket in have else "뼈대"
        print(f"\n══ {bucket}  [{mark}] ══")
        print("  " + body.replace("\n", "\n  "))
    if missing:
        print(f"\n갈래를 못 찾은 시트: {', '.join(missing)}")

    if not args.apply:
        print(f"\n미리보기입니다. 넣으려면 --apply 를 붙이세요 ({len(found)}건)")
        return 0

    for bucket, body in found:
        # 팀 기본 문구(user_id NULL)로 넣는다 — 갈래마다 하나씩.
        row = db.execute(
            select(MessageTemplate).where(
                MessageTemplate.user_id.is_(None),
                MessageTemplate.kind == sourcing_msg.KIND,
                MessageTemplate.name == bucket)
        ).scalars().first()
        if row is None:
            db.add(MessageTemplate(user_id=None, kind=sourcing_msg.KIND,
                                   name=bucket, body=body, is_active=1))
        else:
            row.body = body
            row.is_active = 1
    db.commit()
    print(f"\n{len(found)}건 넣었습니다.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
