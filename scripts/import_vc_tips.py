"""투자사 명단 시트의 `TIPS 운영사 …` 칸을 앱으로 옮긴다.

화면에는 이 칸이 이미 서 있는데(투자사 관리 현황) 값이 비어 있었다. 이 칸은
"이 투자사에 소개해도 되는 기업인가" 를 가르는 값이라(TIPS 운영사는 투자금
1-10억 · 매출 3-20억 기업을 본다), 비어 있으면 딜을 고를 때 시트를 다시 연다.

## 자리가 아니라 **머리글 이름**으로 찾는다
명단 3장에서는 이 칸이 J열이지만, 시트는 사람이 계속 쓰는 문서라 칸이 늘면
자리가 밀린다. `전체 딜소개현황` · `스타트업` 처럼 아예 이 칸이 없는 탭도 있어서
**모든 시트를 훑고 머리글에 `TIPS … 운영사` 가 있는 탭만** 대상으로 삼는다.
없는 탭은 건너뛴다 — 억지로 열 번째 칸을 읽으면 엉뚱한 값이 들어온다.

## 번호가 붙은 줄만 표의 줄이다
`투자사 98명` 시트에는 **칸이 밀린 다른 블록**이 아래로 이어 붙어 있다. 그 줄은
이름 자리에 `담당자2` 같은 라벨이, 회사 자리에 주소가, 그리고 **TIPS 자리에
부서명**(`People & Culture` · `IB3부` · `투자성장본부 / 제1투자그룹`)이 들어온다.
그대로 옮기면 "TIPS 운영사인가" 자리에 부서 이름이 뜬다.

가르는 기준은 짐작이 아니라 시트가 직접 말해 주는 것이어야 한다 —
**`NO` 칸에 번호가 붙은 줄만 표의 줄이다.** 이 시트에서 번호 있는 83줄 /
라벨 블록 61줄로 정확히 갈린다(`import_startup_sheet.py` 가 같은 이유로 같은
기준을 쓴다).

## 사람은 이름+회사로 맞춘다
맞추는 방법은 `import_vc_sheets.py` 에서 **그대로 가져다 쓴다.** 여기서 다시
쓰면 두 벌이 되고, 한쪽만 고쳐지면 그때 매칭되던 사람이 이제 안 맞는다.
동명이인이 있으므로(같은 이름 3명인 경우도 있다) 회사까지 같아야 넣는다 — 애매하면 건너뛰고
끝에 보고한다. 잘못 넣은 값은 화면에서 찾아내기 어렵다.

## 값이 없으면 공란이다
시트가 비어 있는 사람은 건드리지 않는다. 이미 앱에 든 값도 덮지 않는다
(`--overwrite` 로만 덮는다) — 앱에서 고친 내용이 임포트 한 번에 사라지면 안 된다.
그래서 몇 번을 다시 돌려도 결과가 같다.

    python scripts/import_vc_tips.py 파일.xlsx            # 미리보기
    python scripts/import_vc_tips.py 파일.xlsx --apply
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.db import SessionLocal  # noqa: E402
from app.models import VcContact  # noqa: E402
# 맞추는 규칙은 명단 임포터의 것을 그대로 쓴다 — 베껴 두면 두 벌이 되고,
# 한쪽만 고쳐지면 그때 맞던 사람이 이제 안 맞는다.
from scripts.import_vc_sheets import (  # noqa: E402
    _guess_name_column, bare_name, firm_key, header_row, text,
)

FIELD = "tips_note"

# 머리글이 이 칸인가. 원문은 `TIPS 운영사 투자금 1-10억 스타트업매출액 3-20억
# 기업에 주로 투자` 인데, 사람이 문구를 다듬을 수 있어 앞 두 낱말만 본다.
# 메모·주소에도 `TIPS` 가 섞여 있으므로(`TIPS TOWN S6, 6층`) `운영사` 를 함께 건다.
TIPS_HEADER = re.compile(r"TIPS.*운영사", re.IGNORECASE)


def has_row_no(value) -> bool:
    """시트의 `NO` 칸에 번호가 있는가 — 그 줄만 표의 줄이다.

    글자 모양으로 가르려 하면 양쪽으로 틀린다. 번호는 시트가 직접 말해 주는
    사실이라 아래에 이어 붙은 다른 블록과 정확히 갈린다.
    """
    try:
        return float(str(value).strip()) > 0
    except (TypeError, ValueError):
        return False


def find_column(head: dict, want) -> int | None:
    """머리글 이름으로 칸 번호를 찾는다. 자리로 짐작하지 않는다."""
    for label, col in head.items():
        if want(label):
            return col
    return None


def main() -> int:
    ap = argparse.ArgumentParser(description="투자사 명단의 TIPS 칸 가져오기")
    ap.add_argument("path")
    ap.add_argument("--apply", action="store_true", help="실제로 저장")
    ap.add_argument("--overwrite", action="store_true",
                    help="이미 값이 있는 칸도 덮어쓴다")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.path, data_only=True)
    db = SessionLocal()
    people = db.execute(select(VcContact)).scalars().all()

    by_pair, by_name = {}, {}
    for p in people:
        key = bare_name(p.name)
        if not key:
            continue
        by_pair[(key, firm_key(p.firm))] = p
        by_name.setdefault(key, []).append(p)

    filled = kept = 0
    no_column, unmatched, ambiguous, off_table = [], [], [], []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        hr = header_row(ws)
        head = {text(ws.cell(hr, c).value): c for c in range(1, ws.max_column + 1)}
        tips_col = find_column(head, lambda h: bool(TIPS_HEADER.search(h)))
        if tips_col is None:
            no_column.append(sheet)
            continue

        name_col = find_column(head, lambda h: "이름" in h)
        if name_col is None:
            # `투자사 30명` 은 이름 칸의 머리글이 비어 있다.
            name_col = _guess_name_column(ws, hr, head)
        if name_col is None:
            print(f"── {sheet}: 이름 칸을 찾지 못해 건너뜁니다")
            continue
        firm_col = find_column(head, lambda h: h.strip() == "회사")
        no_col = find_column(head, lambda h: h.strip().upper() == "NO")

        found = sheet_filled = sheet_kept = 0
        for r in range(hr + 1, ws.max_row + 1):
            value = text(ws.cell(r, tips_col).value)
            if not value:
                continue        # 값이 없으면 공란이다 — 억지로 채우지 않는다
            # 번호가 없는 줄은 아래에 이어 붙은 다른 블록이다. 그 자리의 값은
            # TIPS 가 아니라 부서명이라 넣으면 안 된다.
            if no_col is not None and not has_row_no(ws.cell(r, no_col).value):
                off_table.append((sheet, r, value))
                continue
            found += 1

            key = bare_name(ws.cell(r, name_col).value)
            firm = firm_key(ws.cell(r, firm_col).value) if firm_col else ""
            person = by_pair.get((key, firm))
            if person is None:
                candidates = by_name.get(key, [])
                if len(candidates) == 1:
                    person = candidates[0]
                elif candidates:
                    # 동명이인인데 회사가 안 맞는다 — 어느 쪽인지 모른 채
                    # 넣으면 엉뚱한 사람에게 붙는다. 넣지 말고 보고한다.
                    ambiguous.append((sheet, r, key, firm))
                    continue
            if person is None:
                unmatched.append((sheet, r, key, firm))
                continue

            if getattr(person, FIELD) and not args.overwrite:
                sheet_kept += 1     # 앱에서 고친 내용을 임포트가 지우면 안 된다
                continue
            setattr(person, FIELD, value)
            sheet_filled += 1

        filled += sheet_filled
        kept += sheet_kept
        print(f"── {sheet} (머리 {hr}행 · {tips_col}번째 칸) ──")
        print(f"   값이 있는 표의 줄 {found}개 · 채움 {sheet_filled} · "
              f"이미 있어 그대로 둠 {sheet_kept}")

    if no_column:
        print(f"\n이 칸이 없는 시트 {len(no_column)}장(건너뜀): "
              + ", ".join(no_column))
    if off_table:
        # 조용히 버리면 몇 줄이 왜 빠졌는지 알 수 없다.
        print(f"\n표 밖의 줄 {len(off_table)}개를 걸렀습니다 "
              f"(번호 없는 줄 — 부서명이 이 자리에 들어와 있다):")
        for sheet, r, value in off_table[:5]:
            print(f"   {sheet} {r}행: {value[:40]}")
    if unmatched:
        print(f"\n앱에서 못 찾은 사람 {len(unmatched)}명: "
              + ", ".join(f"{n}({f})" for _s, _r, n, f in unmatched[:6]))
    if ambiguous:
        print(f"\n동명이인이라 건너뛴 사람 {len(ambiguous)}명 "
              f"(회사가 안 맞아 어느 쪽인지 모른다): "
              + ", ".join(f"{n}({f})" for _s, _r, n, f in ambiguous[:6]))

    print(f"\n채운 칸 {filled}개 · 이미 있어 그대로 둔 칸 {kept}개")
    if args.apply:
        db.commit()
        print("→ 저장했습니다.")
    else:
        db.rollback()
        print("→ 미리보기입니다. 실제로 넣으려면 --apply 를 붙이세요.")
    db.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
